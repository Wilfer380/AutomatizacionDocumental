from __future__ import annotations

import logging
import os
import queue
import threading
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from tkinter import BooleanVar, StringVar, filedialog, messagebox, ttk
import tkinter as tk

from openpyxl import load_workbook

from ..config import DEFAULT_DOSSIER_EXCEL_PATH, DEFAULT_DOSSIER_ROOT_PATH
from ..dossier_models import DossierConfig, DossierExecutionMode, DossierRow, DossierStatus
from ..services.dossier_service import DossierService
from ..services.dossier_validator_service import DossierValidationTreeRow, DossierValidatorService
from ..utils.text import normalize_for_match
from .filter_dropdown import MultiSelectDropdown


logger = logging.getLogger(__name__)

PROGRESS_ANIMATION_MS = 10
QUEUE_POLL_MS = 50


PLACEHOLDER_ROOT_PATH = r"Ejemplo: Q:\GROUPS\CO_MDE_DISENO_DI\01_ORDERS\02_DOCUMENTS_APPROVAL_CERTIFIED"
PLACEHOLDER_EXCEL_PATH = r"Ejemplo: C:\Users\...\Listado equipos SLA COL.xlsx"
PLACEHOLDER_PDF_PATH = "Seleccioná el PDF correspondiente..."
PLACEHOLDER_PHASE1_FOLDER = r"Ejemplo: C:\Users\...\word_excel_pdf_output\PDF_generados"


PALETTE = {
    "background": "#edf3f8",
    "surface": "#ffffff",
    "surface_soft": "#f6f9fc",
    "line": "#d9e3ee",
    "header": "#16395f",
    "header_dark": "#0f2743",
    "accent": "#1e6cf2",
    "accent_dark": "#1559c7",
    "accent_light": "#eaf2ff",
    "text": "#17324d",
    "muted": "#657385",
    "success": "#eaf8ef",
    "success_accent": "#2f9e62",
    "warning": "#fff5e3",
    "warning_accent": "#d79b16",
    "danger": "#fdecec",
    "danger_accent": "#cf4b5b",
    "info": "#eef4ff",
    "info_accent": "#4b72f0",
    "green": "#1bb36a",
}


STATUS_LABELS = {
    DossierStatus.VALID: "Correcto",
    DossierStatus.WARNING: "Advertencia",
    DossierStatus.ERROR: "Error",
    DossierStatus.BLOCKED: "Error",
    DossierStatus.SKIPPED: "Omitido",
    DossierStatus.PLANNED: "Simulado",
    DossierStatus.SIMULATED: "Simulado",
    DossierStatus.COPIED: "Agregado",
    DossierStatus.REPLACED: "Reemplazado",
}


STATUS_TAGS = {
    DossierStatus.VALID: "correct",
    DossierStatus.WARNING: "warning",
    DossierStatus.ERROR: "error",
    DossierStatus.BLOCKED: "error",
    DossierStatus.SKIPPED: "skipped",
    DossierStatus.PLANNED: "correct",
    DossierStatus.SIMULATED: "correct",
    DossierStatus.COPIED: "correct",
    DossierStatus.REPLACED: "correct",
}


STATUS_PRIORITY = {
    DossierStatus.ERROR: 6,
    DossierStatus.BLOCKED: 6,
    DossierStatus.WARNING: 5,
    DossierStatus.COPIED: 4,
    DossierStatus.REPLACED: 4,
    DossierStatus.PLANNED: 3,
    DossierStatus.SIMULATED: 3,
    DossierStatus.VALID: 3,
    DossierStatus.SKIPPED: 2,
}


@dataclass(slots=True)
class DocumentField:
    title: str
    source_var: StringVar
    target_folder: str
    final_name: str
    replacement_note: str


class OperationCancelledError(RuntimeError):
    pass


class DossierPanel(ttk.Frame):
    def __init__(self, parent: tk.Widget) -> None:
        super().__init__(parent, style="App.TFrame")
        self.service = DossierService()
        self.validator = DossierValidatorService()

        self.root_path_var = StringVar(value=PLACEHOLDER_ROOT_PATH)
        self.excel_path_var = StringVar(value=PLACEHOLDER_EXCEL_PATH)
        self.sheet_var = StringVar(value="Hoja1")
        self.cp_column_var = StringVar(value="CP")
        self.serie_column_var = StringVar(value="Serie")

        self.route_status_var = StringVar(value="Sin ruta cargada.")
        self.excel_status_var = StringVar(value="Sin Excel cargado.")

        self.descriptivo_var = StringVar(value=PLACEHOLDER_PDF_PATH)
        self.comunicado_var = StringVar(value=PLACEHOLDER_PDF_PATH)
        self.informe_var = StringVar(value=PLACEHOLDER_PDF_PATH)
        self.phase1_folder_var = StringVar(value=PLACEHOLDER_PHASE1_FOLDER)
        self.use_phase1_var = BooleanVar(value=True)

        self.simulation_var = BooleanVar(value=True)
        self.backup_var = BooleanVar(value=True)
        self.replace_var = BooleanVar(value=True)
        self.continue_var = BooleanVar(value=True)
        self.open_report_var = BooleanVar(value=True)
        self.execution_cp_filter_var = StringVar(value="Todas")
        self.execution_serie_filter_var = StringVar(value="Todas")
        self._selected_cp_filters: list[str] = []
        self._selected_serie_filters: list[str] = []

        self.preview_filter_var = StringVar(value="")
        self.progress_state_var = StringVar(value="Listo para validar CP y Series")
        self.progress_percent_var = StringVar(value="0%")
        self.processed_var = StringVar(value="0 / 0")
        self.correct_var = StringVar(value="0")
        self.warning_var = StringVar(value="0")
        self.error_var = StringVar(value="0")
        self.skipped_var = StringVar(value="0")

        self.report_path_var = StringVar(value="")
        self.log_path_var = StringVar(value="")
        self.last_update_var = StringVar(value="")

        self._operation_queue: queue.Queue = queue.Queue()
        self._operation_worker: threading.Thread | None = None
        self._operation_running = False
        self._current_operation_label = "simulación"
        self._sheet_values: tuple[str, ...] = ("Hoja1",)
        self._cancel_requested = threading.Event()
        self._active_action_button: ttk.Button | None = None
        self._default_button_labels: dict[ttk.Button, str] = {}
        self._excel_rows_cache = []
        self._execution_cp_values: tuple[str, ...] = ()
        self._execution_series_values: tuple[str, ...] = ()
        self._execution_series_by_cp: dict[str, tuple[str, ...]] = {}
        self._validation_rows_cache: list[DossierRow] = []
        self._validation_tree_cache: list[DossierValidationTreeRow] = []
        self._last_summary = None
        self._operation_total = 0

        self.preview_records: list[dict[str, object]] = []
        self.document_fields: list[DocumentField] = []
        self.chart_counts = Counter(
            {
                DossierStatus.VALID: 78,
                DossierStatus.WARNING: 12,
                DossierStatus.ERROR: 8,
                DossierStatus.SKIPPED: 1,
            }
        )

        self._configure_style()
        self._build_ui()
        self._reset_initial_state()
        self.after(QUEUE_POLL_MS, self._poll_operation_queue)

    # ------------------------------------------------------------------ UI
    def _configure_style(self) -> None:
        style = ttk.Style(self)
        self.configure(style="App.TFrame")
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure("App.TFrame", background=PALETTE["background"])
        style.configure("CardBody.TFrame", background=PALETTE["surface"])
        style.configure("App.TLabel", background=PALETTE["surface"], foreground=PALETTE["text"], font=("Segoe UI", 10))
        style.configure("Muted.TLabel", background=PALETTE["surface"], foreground=PALETTE["muted"], font=("Segoe UI", 9))
        style.configure("SectionTitle.TLabel", background=PALETTE["surface"], foreground=PALETTE["accent"], font=("Segoe UI Semibold", 12))
        style.configure("SectionNote.TLabel", background=PALETTE["surface"], foreground=PALETTE["muted"], font=("Segoe UI", 9))
        style.configure("FieldLabel.TLabel", background=PALETTE["surface"], foreground=PALETTE["text"], font=("Segoe UI", 10))
        style.configure("Body.TCheckbutton", background=PALETTE["surface"], foreground=PALETTE["text"], font=("Segoe UI", 10))
        style.configure("App.TEntry", padding=7)
        style.configure("App.TCombobox", padding=5)
        style.configure("App.Treeview", background=PALETTE["surface"], fieldbackground=PALETTE["surface"], foreground=PALETTE["text"], rowheight=30)
        style.configure("App.Treeview.Heading", background=PALETTE["surface_soft"], foreground=PALETTE["text"], font=("Segoe UI Semibold", 10), padding=7)
        style.map("App.Treeview", background=[("selected", PALETTE["accent_light"])], foreground=[("selected", PALETTE["text"])])
        style.configure("Primary.TButton", padding=(18, 11), font=("Segoe UI Semibold", 10), background=PALETTE["accent"], foreground="white")
        style.map("Primary.TButton", background=[("active", PALETTE["accent_dark"]), ("pressed", PALETTE["accent_dark"])])
        style.configure("Green.TButton", padding=(18, 11), font=("Segoe UI Semibold", 10), background=PALETTE["green"], foreground="white")
        style.map("Green.TButton", background=[("active", "#19a35e"), ("pressed", "#19a35e")])
        style.configure("Gray.TButton", padding=(18, 11), font=("Segoe UI Semibold", 10), background="#e7eef7", foreground=PALETTE["text"])
        style.map("Gray.TButton", background=[("active", "#d7e2f0"), ("pressed", "#d7e2f0")])
        style.configure("Ghost.TButton", padding=(12, 8), font=("Segoe UI Semibold", 10), background=PALETTE["surface"], foreground=PALETTE["text"])
        style.map("Ghost.TButton", background=[("active", PALETTE["surface_soft"]), ("pressed", PALETTE["surface_soft"])])
        style.configure("Tab.TButton", padding=(0, 0), background=PALETTE["background"], foreground=PALETTE["text"], font=("Segoe UI Semibold", 10))
        style.configure("StatusBar.TLabel", background=PALETTE["accent_light"], foreground=PALETTE["text"], font=("Segoe UI Semibold", 9))
        style.configure("InfoBox.TLabel", background=PALETTE["info"], foreground=PALETTE["text"], font=("Segoe UI", 9))

    def _build_ui(self) -> None:
        viewport = ttk.Frame(self, style="App.TFrame")
        viewport.pack(fill="both", expand=True)
        viewport.columnconfigure(0, weight=1)
        viewport.rowconfigure(0, weight=1)

        self.canvas = tk.Canvas(viewport, bg=PALETTE["background"], highlightthickness=0)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(viewport, orient="vertical", command=self.canvas.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.canvas.configure(yscrollcommand=scrollbar.set)

        self.content = ttk.Frame(self.canvas, style="App.TFrame", padding=(18, 14, 18, 14))
        self.content_window = self.canvas.create_window((0, 0), window=self.content, anchor="nw")
        self.content.bind("<Configure>", self._sync_scroll_region)
        self.canvas.bind("<Configure>", self._sync_content_width)
        self._bind_mousewheel(self.canvas)

        outer = self.content
        outer.columnconfigure(0, weight=47)
        outer.columnconfigure(1, weight=53)
        outer.rowconfigure(0, weight=1)

        self.left = ttk.Frame(outer, style="App.TFrame")
        self.right = ttk.Frame(outer, style="App.TFrame")
        self.left.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        self.right.grid(row=0, column=1, sticky="nsew")
        self.left.columnconfigure(0, weight=1)
        self.right.columnconfigure(0, weight=1)
        self.right.rowconfigure(0, weight=2)
        self.right.rowconfigure(1, weight=1)
        self.right.rowconfigure(2, weight=1)

        self._build_section_1(self.left)
        self._build_section_2(self.left)
        self._build_section_3(self.left)
        self._build_section_4(self.left)
        self._build_section_5(self.left)

        self._build_section_6(self.right)
        self._build_section_7(self.right)
        self._build_section_8(self.right)

    def _sync_scroll_region(self, _event: tk.Event) -> None:
        if hasattr(self, "canvas"):
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _sync_content_width(self, event: tk.Event) -> None:
        if hasattr(self, "content_window"):
            self.canvas.itemconfigure(self.content_window, width=event.width)

    def _bind_mousewheel(self, widget: tk.Widget) -> None:
        widget.bind_all("<MouseWheel>", self._on_mousewheel, add="+")
        widget.bind_all("<Button-4>", self._on_mousewheel, add="+")
        widget.bind_all("<Button-5>", self._on_mousewheel, add="+")

    def _on_mousewheel(self, event: tk.Event) -> None:
        if not hasattr(self, "canvas"):
            return
        if getattr(event, "num", None) == 4:
            self.canvas.yview_scroll(-1, "units")
        elif getattr(event, "num", None) == 5:
            self.canvas.yview_scroll(1, "units")
        else:
            delta = int(-1 * (event.delta / 120))
            self.canvas.yview_scroll(delta, "units")

    def _section_card(self, parent: ttk.Frame, title: str, accent: str, subtitle: str | None = None) -> tuple[tk.Frame, ttk.Frame]:
        outer = tk.Frame(parent, bg=accent, highlightthickness=0)
        inner = tk.Frame(outer, bg=PALETTE["surface"], highlightthickness=0)
        inner.pack(fill="both", expand=True, padx=1, pady=1)

        header = tk.Frame(inner, bg=PALETTE["surface"])
        header.pack(fill="x", padx=16, pady=(14, 10))
        tk.Label(header, text=title, bg=PALETTE["surface"], fg=PALETTE["accent"], font=("Segoe UI Semibold", 12)).pack(anchor="w")
        if subtitle:
            tk.Label(header, text=subtitle, bg=PALETTE["surface"], fg=PALETTE["muted"], font=("Segoe UI", 9), wraplength=420, justify="left").pack(anchor="w", pady=(2, 0))

        body = ttk.Frame(inner, style="CardBody.TFrame", padding=(16, 0, 16, 16))
        body.pack(fill="both", expand=True)
        return outer, body

    def _entry_row(self, parent: ttk.Frame, row: int, label: str, variable: StringVar, browse_command, browse_text: str = "Examinar") -> None:
        ttk.Label(parent, text=label, style="FieldLabel.TLabel").grid(row=row, column=0, sticky="w", pady=6)
        ttk.Entry(parent, textvariable=variable, style="App.TEntry").grid(row=row, column=1, sticky="ew", padx=8, pady=6)
        ttk.Button(parent, text=browse_text, style="Ghost.TButton", command=browse_command).grid(row=row, column=2, sticky="e", pady=6)

    def _combo_field(self, parent: ttk.Frame, column: int, label: str, variable: StringVar, values: tuple[str, ...]) -> ttk.Combobox:
        cell = ttk.Frame(parent, style="CardBody.TFrame")
        cell.grid(row=0, column=column, sticky="ew", padx=(0, 10) if column < 2 else 0)
        ttk.Label(cell, text=label, style="FieldLabel.TLabel").pack(anchor="w")
        combo = ttk.Combobox(cell, textvariable=variable, values=values, state="readonly", style="App.TCombobox")
        combo.pack(fill="x", pady=(4, 0))
        return combo

    def _small_badge(self, parent: tk.Widget, text: str, fg: str, bg: str = PALETTE["surface"]) -> tk.Label:
        return tk.Label(parent, text=text, bg=bg, fg=fg, font=("Segoe UI Semibold", 9))

    def _action_button(self, parent: ttk.Frame, row: int, column: int, text: str, style: str, command) -> ttk.Button:
        button = ttk.Button(parent, text=text, style=style, command=command)
        button.grid(row=row, column=column, sticky="ew", padx=5, pady=5)
        return button

    def _count_card(self, parent: ttk.Frame, column: int, title: str, variable: StringVar, accent: str) -> None:
        outer = tk.Frame(parent, bg=accent, highlightthickness=0)
        outer.grid(row=0, column=column, sticky="nsew", padx=6)
        inner = tk.Frame(outer, bg=PALETTE["surface"], padx=14, pady=10)
        inner.pack(fill="both", expand=True, padx=1, pady=1)
        tk.Label(inner, text=title, bg=PALETTE["surface"], fg=accent, font=("Segoe UI", 9)).pack(anchor="w")
        tk.Label(inner, textvariable=variable, bg=PALETTE["surface"], fg=PALETTE["text"], font=("Segoe UI Semibold", 18)).pack(anchor="w", pady=(3, 0))

    def _readonly_path_field(self, parent: ttk.Frame, row: int, label: str, variable: StringVar, button_text: str, command) -> None:
        ttk.Label(parent, text=label, style="FieldLabel.TLabel").grid(row=row, column=0, sticky="w", pady=6)
        ttk.Entry(parent, textvariable=variable, style="App.TEntry", state="readonly").grid(row=row, column=1, sticky="ew", pady=6, padx=8)
        ttk.Button(parent, text=button_text, style="Ghost.TButton", command=command).grid(row=row, column=2, sticky="e", pady=6)

    # ------------------------------------------------------------------ Sections
    def _build_section_1(self, parent: ttk.Frame) -> None:
        card, body = self._section_card(parent, "1. Ruta documental raíz", PALETTE["info_accent"])
        card.grid(row=0, column=0, sticky="ew")
        body.columnconfigure(1, weight=1)
        self._entry_row(body, 0, "Ruta raíz:", self.root_path_var, self._browse_root)

        buttons = ttk.Frame(body, style="CardBody.TFrame")
        buttons.grid(row=0, column=3, sticky="e", padx=(8, 0))
        ttk.Button(buttons, text="Validar ruta", style="Ghost.TButton", command=self.validate_root_path).pack(anchor="e")

        status = tk.Frame(body, bg=PALETTE["accent_light"], highlightthickness=1, highlightbackground=PALETTE["line"])
        status.grid(row=1, column=0, columnspan=4, sticky="ew", pady=(10, 0))
        tk.Label(status, text="✔", bg=PALETTE["accent_light"], fg=PALETTE["success_accent"], font=("Segoe UI", 12)).pack(side="left", padx=(12, 4), pady=10)
        tk.Label(status, textvariable=self.route_status_var, bg=PALETTE["accent_light"], fg=PALETTE["success_accent"], font=("Segoe UI Semibold", 9)).pack(side="left", padx=(0, 12), pady=10)

    def _build_section_2(self, parent: ttk.Frame) -> None:
        card, body = self._section_card(parent, "2. Excel de CP y Series", PALETTE["success_accent"])
        card.grid(row=1, column=0, sticky="ew", pady=(12, 0))
        body.columnconfigure(1, weight=1)
        self._entry_row(body, 0, "Archivo Excel:", self.excel_path_var, self._browse_excel)

        fields = ttk.Frame(body, style="CardBody.TFrame")
        fields.grid(row=1, column=0, columnspan=4, sticky="ew", pady=(8, 0))
        fields.columnconfigure(0, weight=1)
        fields.columnconfigure(1, weight=1)
        fields.columnconfigure(2, weight=1)
        self.sheet_combo = self._combo_field(fields, 0, "Hoja:", self.sheet_var, self._sheet_values)
        self.sheet_combo.bind("<<ComboboxSelected>>", lambda _event: self._refresh_excel_metadata())
        self._combo_field(fields, 1, "Columna CP:", self.cp_column_var, ("CP",))
        self._combo_field(fields, 2, "Columna Serie:", self.serie_column_var, ("Serie",))

        filters = ttk.Frame(body, style="CardBody.TFrame")
        filters.grid(row=2, column=0, columnspan=4, sticky="ew", pady=(10, 0))
        filters.columnconfigure(1, weight=1)
        filters.columnconfigure(3, weight=1)

        ttk.Label(filters, text="Filtro CP:", style="FieldLabel.TLabel").grid(row=0, column=0, sticky="w", pady=6)
        self.execution_cp_dropdown = MultiSelectDropdown(filters, title="Filtro CP", on_selection_changed=self._on_cp_filter_changed)
        self.execution_cp_dropdown.grid(row=0, column=1, sticky="ew", padx=(8, 16), pady=6)

        ttk.Label(filters, text="Filtro Serie:", style="FieldLabel.TLabel").grid(row=0, column=2, sticky="w", pady=6)
        self.execution_serie_dropdown = MultiSelectDropdown(filters, title="Filtro Serie", on_selection_changed=self._on_serie_filter_changed)
        self.execution_serie_dropdown.grid(row=0, column=3, sticky="ew", padx=(8, 8), pady=6)

        ttk.Button(filters, text="Limpiar", style="Ghost.TButton", command=self._clear_execution_filters).grid(row=0, column=4, sticky="e", pady=6)
        tk.Label(
            filters,
            text="Seleccioná una o varias CP y, si querés, una o varias series. Si dejás 'Todas', procesa todo.",
            bg=PALETTE["surface"],
            fg=PALETTE["muted"],
            font=("Segoe UI", 8),
        ).grid(row=1, column=0, columnspan=5, sticky="w", pady=(2, 0))

        status = tk.Frame(body, bg=PALETTE["accent_light"], highlightthickness=1, highlightbackground=PALETTE["line"])
        status.grid(row=3, column=0, columnspan=4, sticky="ew", pady=(10, 0))
        tk.Label(status, text="✔", bg=PALETTE["accent_light"], fg=PALETTE["success_accent"], font=("Segoe UI", 12)).pack(side="left", padx=(12, 4), pady=10)
        tk.Label(status, textvariable=self.excel_status_var, bg=PALETTE["accent_light"], fg=PALETTE["success_accent"], font=("Segoe UI Semibold", 9)).pack(side="left", padx=(0, 12), pady=10)

    def _build_section_3(self, parent: ttk.Frame) -> None:
        card, body = self._section_card(parent, "3. Documentos a distribuir", PALETTE["accent"])
        card.grid(row=2, column=0, sticky="ew", pady=(12, 0))
        body.columnconfigure(1, weight=1)

        self.document_fields = [
            DocumentField(
                title="Descriptivo de pintura (5.2):",
                source_var=self.descriptivo_var,
                target_folder="5 Procedimiento de fabricación",
                final_name="5.2 Descriptivo de pintura.pdf",
                replacement_note="Reemplaza 5.2 Procedimiento de Aplicación de Pintura.pdf.",
            ),
            DocumentField(
                title="Comunicado técnico (7.1):",
                source_var=self.comunicado_var,
                target_folder="7 Ensayos",
                final_name="7.1 Comunicado técnico - Resultados de adherencia.pdf",
                replacement_note="Se copia en carpeta 7 Ensayos.",
            ),
            DocumentField(
                title="Informe laboratorio (7.2):",
                source_var=self.informe_var,
                target_folder="7 Ensayos",
                final_name="7.2 Informe de Ensayo Laboratorio - prueba adherencia.pdf",
                replacement_note="Se copia en carpeta 7 Ensayos.",
            ),
        ]

        for index, field in enumerate(self.document_fields):
            row = ttk.Frame(body, style="CardBody.TFrame")
            row.grid(row=index, column=0, columnspan=4, sticky="ew", pady=(0, 12))
            row.columnconfigure(1, weight=1)
            ttk.Label(row, text=field.title, style="FieldLabel.TLabel").grid(row=0, column=0, sticky="w")
            ttk.Entry(row, textvariable=field.source_var, style="App.TEntry").grid(row=0, column=1, sticky="ew", padx=8)
            ttk.Button(row, text="Examinar", style="Ghost.TButton", command=lambda var=field.source_var: self._browse_pdf(var)).grid(row=0, column=2, sticky="e")
            tk.Label(
                row,
                text=f"Destino: {field.target_folder} · {field.final_name} · {field.replacement_note}",
                bg=PALETTE["surface"],
                fg=PALETTE["muted"],
                font=("Segoe UI", 8),
                wraplength=560,
                justify="left",
            ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(4, 0))

        self.use_phase1_check = ttk.Checkbutton(
            body,
            text="Usar PDFs generados por Fase 1 para Certificado de Trazabilidad (carpeta 6)",
            variable=self.use_phase1_var,
            style="Body.TCheckbutton",
            command=self._toggle_phase1_folder_state,
        )
        self.use_phase1_check.grid(row=3, column=0, columnspan=4, sticky="w", pady=(2, 8))

        phase1_row = ttk.Frame(body, style="CardBody.TFrame")
        phase1_row.grid(row=4, column=0, columnspan=4, sticky="ew")
        phase1_row.columnconfigure(1, weight=1)
        ttk.Label(phase1_row, text="Carpeta de PDFs generados:", style="FieldLabel.TLabel").grid(row=0, column=0, sticky="w", pady=6)
        ttk.Entry(phase1_row, textvariable=self.phase1_folder_var, style="App.TEntry").grid(row=0, column=1, sticky="ew", padx=8, pady=6)
        ttk.Button(phase1_row, text="Examinar", style="Ghost.TButton", command=self._browse_phase1_folder).grid(row=0, column=2, sticky="e", pady=6)

    def _build_section_4(self, parent: ttk.Frame) -> None:
        card, body = self._section_card(parent, "4. Opciones", PALETTE["warning_accent"])
        card.grid(row=3, column=0, sticky="ew", pady=(12, 0))
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)

        left = ttk.Frame(body, style="CardBody.TFrame")
        left.grid(row=0, column=0, sticky="nw")
        for text, variable in (
            ("Modo simulación (Dry Run)", self.simulation_var),
            ("Crear respaldo antes de reemplazar", self.backup_var),
            ("Reemplazar si ya existe", self.replace_var),
            ("Continuar aunque haya advertencias", self.continue_var),
            ("Abrir reporte al finalizar", self.open_report_var),
        ):
            ttk.Checkbutton(left, text=text, variable=variable, style="Body.TCheckbutton").pack(anchor="w", pady=4)

        right = tk.Frame(body, bg=PALETTE["info"], highlightthickness=1, highlightbackground=PALETTE["line"])
        right.grid(row=0, column=1, sticky="nsew", padx=(12, 0))
        tk.Label(right, text="Modo simulación activo", bg=PALETTE["info"], fg=PALETTE["header"], font=("Segoe UI Semibold", 10)).pack(anchor="w", padx=12, pady=(10, 2))
        tk.Label(
            right,
            text="El sistema validará todas las rutas, carpetas y archivos, pero NO copiará, reemplazará ni moverá ningún archivo.",
            bg=PALETTE["info"],
            fg=PALETTE["text"],
            font=("Segoe UI", 9),
            wraplength=310,
            justify="left",
        ).pack(anchor="w", padx=12, pady=(0, 10))

    def _build_section_5(self, parent: ttk.Frame) -> None:
        card, body = self._section_card(parent, "5. Acciones", PALETTE["accent"])
        card.grid(row=4, column=0, sticky="ew", pady=(12, 0))
        for column in range(4):
            body.columnconfigure(column, weight=1)

        self.validate_button = self._action_button(body, 0, 0, "Validar CP y Series", "Primary.TButton", self.validate_cp_and_series)
        self.simulate_button = self._action_button(body, 0, 1, "Simular distribución", "Primary.TButton", self.simulate_distribution)
        self.real_button = self._action_button(body, 0, 2, "Ejecutar distribución real", "Green.TButton", self.execute_real_distribution)
        self.cancel_button = self._action_button(body, 0, 3, "Cancelar", "Gray.TButton", self.cancel_operation)

        utilities = ttk.Frame(body, style="CardBody.TFrame")
        utilities.grid(row=1, column=0, columnspan=4, sticky="ew", pady=(10, 0))
        utilities.columnconfigure(0, weight=1)
        utilities.columnconfigure(1, weight=1)
        utilities.columnconfigure(2, weight=1)
        ttk.Button(utilities, text="Abrir reporte", style="Ghost.TButton", command=self.open_report).grid(row=0, column=0, sticky="ew", padx=(0, 6))
        ttk.Button(utilities, text="Abrir log", style="Ghost.TButton", command=self.open_log).grid(row=0, column=1, sticky="ew", padx=(6, 6))
        ttk.Button(utilities, text="Limpiar todo", style="Ghost.TButton", command=self.clear_all_fields).grid(row=0, column=2, sticky="ew", padx=(6, 0))

    def _build_section_6(self, parent: ttk.Frame) -> None:
        card, body = self._section_card(parent, "6. Vista previa de la distribución", PALETTE["accent"])
        card.grid(row=0, column=0, sticky="nsew")
        body.columnconfigure(0, weight=1)
        body.rowconfigure(1, weight=1)

        header = ttk.Frame(body, style="CardBody.TFrame")
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=1)

        filter_area = ttk.Frame(header, style="CardBody.TFrame")
        filter_area.grid(row=0, column=1, sticky="e")
        ttk.Label(filter_area, text="Filtrar:", style="FieldLabel.TLabel").pack(side="left", padx=(0, 8))
        filter_entry = ttk.Entry(filter_area, textvariable=self.preview_filter_var, width=24, style="App.TEntry")
        filter_entry.pack(side="left")
        filter_entry.bind("<KeyRelease>", lambda _event: self._refresh_preview_tree())
        ttk.Button(filter_area, text="Limpiar", style="Ghost.TButton", command=self._clear_preview_filter).pack(side="left", padx=(8, 0))

        table_frame = ttk.Frame(body, style="CardBody.TFrame")
        table_frame.grid(row=1, column=0, sticky="nsew", pady=(12, 0))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        columns = (
            ("fila_excel", "Fila Excel", 70, "center"),
            ("cp", "CP", 120, "w"),
            ("serie", "Serie", 122, "w"),
            ("carpeta_cp", "Carpeta CP encontrada", 185, "w"),
            ("serie_planos", "Serie encontrada", 118, "center"),
            ("carpeta_5", "Carpeta 5", 90, "center"),
            ("carpeta_6", "Carpeta 6", 90, "center"),
            ("carpeta_7", "Carpeta 7", 90, "center"),
            ("accion", "Acción prevista", 132, "w"),
            ("estado", "Estado", 108, "center"),
            ("observacion", "Observación", 210, "w"),
        )
        self.preview_tree = ttk.Treeview(table_frame, columns=[item[0] for item in columns], show="headings", style="App.Treeview", selectmode="browse")
        for col_id, heading, width, anchor in columns:
            self.preview_tree.heading(col_id, text=heading)
            self.preview_tree.column(col_id, width=width, anchor=anchor, stretch=col_id == "observacion")
        self.preview_tree.grid(row=0, column=0, sticky="nsew")
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.preview_tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.preview_tree.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        self.preview_tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self.preview_tree.tag_configure("correct", background=PALETTE["success"])
        self.preview_tree.tag_configure("warning", background=PALETTE["warning"])
        self.preview_tree.tag_configure("error", background=PALETTE["danger"])
        self.preview_tree.tag_configure("skipped", background=PALETTE["accent_light"])

    def _build_section_7(self, parent: ttk.Frame) -> None:
        card, body = self._section_card(parent, "7. Progreso", PALETTE["success_accent"])
        card.grid(row=1, column=0, sticky="ew", pady=(12, 0))
        body.columnconfigure(0, weight=1)

        top = ttk.Frame(body, style="CardBody.TFrame")
        top.grid(row=0, column=0, sticky="ew")
        top.columnconfigure(0, weight=1)

        line1 = ttk.Frame(top, style="CardBody.TFrame")
        line1.grid(row=0, column=0, sticky="ew")
        ttk.Label(line1, text="Estado actual:", style="FieldLabel.TLabel").pack(side="left")
        ttk.Label(line1, textvariable=self.progress_state_var, style="FieldLabel.TLabel").pack(side="left", padx=(8, 0))
        ttk.Label(line1, textvariable=self.progress_percent_var, style="FieldLabel.TLabel").pack(side="right")

        self.progress = ttk.Progressbar(body, mode="determinate", maximum=100)
        self.progress.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        ttk.Label(body, textvariable=self.processed_var, style="Muted.TLabel").grid(row=2, column=0, sticky="w", pady=(6, 0))

        metrics = ttk.Frame(body, style="CardBody.TFrame")
        metrics.grid(row=3, column=0, sticky="ew", pady=(14, 0))
        for index in range(4):
            metrics.columnconfigure(index, weight=1)
        self._count_card(metrics, 0, "Correctos", self.correct_var, PALETTE["success_accent"])
        self._count_card(metrics, 1, "Advertencias", self.warning_var, PALETTE["warning_accent"])
        self._count_card(metrics, 2, "Errores", self.error_var, PALETTE["danger_accent"])
        self._count_card(metrics, 3, "Omitidos", self.skipped_var, PALETTE["info_accent"])

    def _build_section_8(self, parent: ttk.Frame) -> None:
        card, body = self._section_card(parent, "8. Resumen y trazabilidad", PALETTE["danger_accent"])
        card.grid(row=2, column=0, sticky="ew", pady=(12, 0))
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)

        left = ttk.Frame(body, style="CardBody.TFrame")
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        left.columnconfigure(1, weight=1)
        self._readonly_path_field(left, 0, "Reporte:", self.report_path_var, "Abrir reporte", self.open_report)
        self._readonly_path_field(left, 1, "Log técnico:", self.log_path_var, "Abrir log", self.open_log)
        ttk.Label(left, text="Última actualización:", style="FieldLabel.TLabel").grid(row=2, column=0, sticky="w", pady=(6, 0))
        ttk.Label(left, textvariable=self.last_update_var, style="FieldLabel.TLabel").grid(row=2, column=1, sticky="w", pady=(6, 0), padx=8)

        right = tk.Frame(body, bg=PALETTE["surface"], highlightthickness=1, highlightbackground=PALETTE["line"])
        right.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        right.columnconfigure(0, weight=1)
        right.columnconfigure(1, weight=1)

        self.chart_canvas = tk.Canvas(right, width=250, height=190, bg=PALETTE["surface"], highlightthickness=0)
        self.chart_canvas.grid(row=0, column=0, rowspan=2, sticky="n", padx=(10, 6), pady=10)
        self.legend_frame = tk.Frame(right, bg=PALETTE["surface"])
        self.legend_frame.grid(row=0, column=1, sticky="n", padx=(0, 10), pady=10)

    # ------------------------------------------------------------------ Helpers / State
    def _reset_initial_state(self) -> None:
        self.preview_records = []
        self.chart_counts = Counter({DossierStatus.VALID: 0, DossierStatus.WARNING: 0, DossierStatus.ERROR: 0, DossierStatus.SKIPPED: 0})
        self.progress["value"] = 0
        self.progress_percent_var.set("0%")
        self.processed_var.set("0 / 0")
        self.correct_var.set("0")
        self.warning_var.set("0")
        self.error_var.set("0")
        self.skipped_var.set("0")
        self.progress_state_var.set("Listo para validar CP y Series")
        self.report_path_var.set("")
        self.log_path_var.set("")
        self.last_update_var.set("")
        self._refresh_preview_tree()
        self._refresh_chart()

    def _refresh_chart(self) -> None:
        if not hasattr(self, "chart_canvas"):
            return
        self.chart_canvas.delete("all")
        for child in self.legend_frame.winfo_children():
            child.destroy()

        items = [
            ("Correctos", self.chart_counts.get(DossierStatus.VALID, 0), PALETTE["success_accent"]),
            ("Advertencias", self.chart_counts.get(DossierStatus.WARNING, 0), PALETTE["warning_accent"]),
            ("Errores", self.chart_counts.get(DossierStatus.ERROR, 0), PALETTE["danger_accent"]),
            ("Omitidos", self.chart_counts.get(DossierStatus.SKIPPED, 0), PALETTE["info_accent"]),
        ]
        total = sum(value for _label, value, _color in items)
        if total <= 0:
            self.chart_canvas.create_oval(60, 28, 180, 148, fill=PALETTE["accent_light"], outline=PALETTE["line"])
            self.chart_canvas.create_text(120, 88, text="Sin datos", fill=PALETTE["muted"], font=("Segoe UI Semibold", 10))
        else:
            start = 90
            for label, value, color in items:
                if value <= 0:
                    continue
                extent = -360 * value / total
                self.chart_canvas.create_arc(36, 20, 204, 188, start=start, extent=extent, fill=color, outline=PALETTE["surface"])
                mid = start + extent / 2
                x = 120 + 58 * self._cos(mid)
                y = 104 - 58 * self._sin(mid)
                self.chart_canvas.create_text(x, y, text=f"{value / total * 100:.1f}%", fill="white", font=("Segoe UI Semibold", 9))
                start += extent
            self.chart_canvas.create_oval(74, 58, 166, 150, fill=PALETTE["surface"], outline=PALETTE["surface"])
            self.chart_canvas.create_text(120, 99, text=f"{total}", fill=PALETTE["text"], font=("Segoe UI Semibold", 18))
            self.chart_canvas.create_text(120, 122, text="registros", fill=PALETTE["muted"], font=("Segoe UI", 9))

        for label, value, color in items:
            row = tk.Frame(self.legend_frame, bg=PALETTE["surface"])
            row.pack(anchor="w", fill="x", pady=4)
            marker = tk.Canvas(row, width=12, height=12, bg=PALETTE["surface"], highlightthickness=0)
            marker.pack(side="left")
            marker.create_oval(1, 1, 11, 11, fill=color, outline=color)
            tk.Label(row, text=f"{label} ({value})", bg=PALETTE["surface"], fg=PALETTE["text"], font=("Segoe UI", 9)).pack(side="left", padx=(8, 0))

    @staticmethod
    def _cos(angle_degrees: float) -> float:
        import math

        return math.cos(math.radians(angle_degrees))

    @staticmethod
    def _sin(angle_degrees: float) -> float:
        import math

        return math.sin(math.radians(angle_degrees))

    def _toggle_phase1_folder_state(self) -> None:
        if not self.use_phase1_var.get():
            self.phase1_folder_var.set(PLACEHOLDER_PHASE1_FOLDER)

    @staticmethod
    def _is_placeholder_value(value: str) -> bool:
        raw = (value or "").strip()
        return raw in {
            PLACEHOLDER_ROOT_PATH,
            PLACEHOLDER_EXCEL_PATH,
            PLACEHOLDER_PDF_PATH,
            PLACEHOLDER_PHASE1_FOLDER,
        }

    def _read_input(self, variable: StringVar) -> str:
        raw = (variable.get() or "").strip()
        return "" if self._is_placeholder_value(raw) else raw

    # ------------------------------------------------------------------ Browse / validate
    def _browse_root(self) -> None:
        initial_dir = self._read_input(self.root_path_var) or str(Path.home())
        path = filedialog.askdirectory(initialdir=initial_dir)
        if path:
            self.root_path_var.set(path)
            self.validate_root_path()

    def _browse_excel(self) -> None:
        excel_value = self._read_input(self.excel_path_var)
        initial = Path(excel_value).parent if excel_value else Path.home()
        path = filedialog.askopenfilename(initialdir=str(initial), filetypes=[("Excel", "*.xlsx")])
        if path:
            self.excel_path_var.set(path)
            self._refresh_excel_metadata()

    def _browse_pdf(self, variable: StringVar) -> None:
        path = filedialog.askopenfilename(filetypes=[("PDF", "*.pdf")])
        if path:
            variable.set(path)

    def _browse_phase1_folder(self) -> None:
        initial_dir = self._read_input(self.phase1_folder_var) or str(Path.home())
        path = filedialog.askdirectory(initialdir=initial_dir)
        if path:
            self.phase1_folder_var.set(path)

    def validate_root_path(self) -> None:
        root_value = self._read_input(self.root_path_var)
        if not root_value:
            self.route_status_var.set("Sin ruta cargada.")
            return
        root = Path(root_value)
        if root.exists():
            self.route_status_var.set("Ruta válida.")
        else:
            self.route_status_var.set("Ruta no encontrada.")

    def _refresh_excel_metadata(self) -> None:
        excel_value = self._read_input(self.excel_path_var)
        if not excel_value:
            self._excel_rows_cache = []
            self._populate_execution_filters([])
            self.excel_status_var.set("Sin Excel cargado.")
            return
        excel_path = Path(excel_value)
        if not excel_path.exists():
            self._excel_rows_cache = []
            self._populate_execution_filters([])
            self.excel_status_var.set("Excel no encontrado.")
            return

        try:
            workbook = load_workbook(excel_path, read_only=True, data_only=True)
            try:
                sheet_names = tuple(workbook.sheetnames)
            finally:
                workbook.close()

            if sheet_names:
                self._sheet_values = sheet_names
                if hasattr(self, "sheet_combo"):
                    self.sheet_combo.configure(values=self._sheet_values)
                self.sheet_var.set(self.sheet_var.get() if self.sheet_var.get() in sheet_names else sheet_names[0])
            config = self._build_config(simulation_only=True)
            _, rows = self.validator.load_rows(config)
            self._excel_rows_cache = rows
            self._populate_execution_filters(rows)
            self.excel_status_var.set(f"Excel cargado correctamente. {len(rows)} registro(s) encontrados.")
        except Exception as exc:  # pragma: no cover - UI feedback
            logger.exception("Unable to inspect workbook metadata")
            self._excel_rows_cache = []
            self._populate_execution_filters([])
            self.excel_status_var.set(f"Error al cargar Excel: {exc}")

    def _normalize_execution_filter(self, value: str) -> str:
        raw = (value or "").strip()
        return "" if raw.lower() == "todas" else raw

    def _get_available_series_for_selected_cps(self) -> tuple[str, ...]:
        if not self._selected_cp_filters:
            return self._execution_series_values

        available: list[str] = []
        seen: set[str] = set()
        for cp in self._selected_cp_filters:
            for serie in self._execution_series_by_cp.get(cp, ()): 
                if serie and serie not in seen:
                    seen.add(serie)
                    available.append(serie)
        return tuple(available)

    def _populate_execution_filters(self, rows: list[DossierRow]) -> None:
        cp_values: list[str] = []
        all_series: list[str] = []
        series_by_cp: dict[str, list[str]] = {}
        seen_cps: set[str] = set()
        seen_series: set[str] = set()

        for row in rows:
            cp_value = str(row.cp or "").strip()
            serie_value = str(row.serie or "").strip()
            if cp_value and cp_value not in seen_cps:
                seen_cps.add(cp_value)
                cp_values.append(cp_value)
            if serie_value and serie_value not in seen_series:
                seen_series.add(serie_value)
                all_series.append(serie_value)
            if cp_value and serie_value:
                series_by_cp.setdefault(cp_value, [])
                if serie_value not in series_by_cp[cp_value]:
                    series_by_cp[cp_value].append(serie_value)

        self._execution_cp_values = tuple(cp_values)
        self._execution_series_by_cp = {cp: tuple(values) for cp, values in series_by_cp.items()}
        self._execution_series_values = tuple(all_series)
        self._selected_cp_filters = [cp for cp in self._selected_cp_filters if cp in self._execution_cp_values]
        self._refresh_execution_filter_widgets()

    def _refresh_execution_filter_widgets(self) -> None:
        available_series = self._get_available_series_for_selected_cps()
        self._selected_serie_filters = [serie for serie in self._selected_serie_filters if serie in available_series]
        if hasattr(self, "execution_cp_dropdown"):
            self.execution_cp_dropdown.set_options(self._execution_cp_values)
            self.execution_cp_dropdown.set_selected(self._selected_cp_filters)
        if hasattr(self, "execution_serie_dropdown"):
            self.execution_serie_dropdown.set_options(available_series)
            self.execution_serie_dropdown.set_selected(self._selected_serie_filters)
            self.execution_serie_dropdown.set_enabled(bool(available_series))

    def _on_cp_filter_changed(self, selected: list[str]) -> None:
        self._selected_cp_filters = list(selected)
        self._refresh_execution_filter_widgets()

    def _on_serie_filter_changed(self, selected: list[str]) -> None:
        self._selected_serie_filters = list(selected)
        self._refresh_execution_filter_widgets()

    def _clear_execution_filters(self) -> None:
        self._selected_cp_filters = []
        self._selected_serie_filters = []
        self._refresh_execution_filter_widgets()

    def clear_all_fields(self) -> None:
        if self._operation_running:
            messagebox.showinfo("Fase 2", "No se puede limpiar mientras hay una operación en curso.")
            return

        self.root_path_var.set(PLACEHOLDER_ROOT_PATH)
        self.excel_path_var.set(PLACEHOLDER_EXCEL_PATH)
        self.sheet_var.set("Hoja1")
        self.cp_column_var.set("CP")
        self.serie_column_var.set("Serie")
        self.descriptivo_var.set(PLACEHOLDER_PDF_PATH)
        self.comunicado_var.set(PLACEHOLDER_PDF_PATH)
        self.informe_var.set(PLACEHOLDER_PDF_PATH)
        self.phase1_folder_var.set(PLACEHOLDER_PHASE1_FOLDER)

        self.use_phase1_var.set(True)
        self.simulation_var.set(True)
        self.backup_var.set(True)
        self.replace_var.set(True)
        self.continue_var.set(True)
        self.open_report_var.set(True)

        self.route_status_var.set("Sin ruta cargada.")
        self.excel_status_var.set("Sin Excel cargado.")
        self.preview_filter_var.set("")

        self._sheet_values = ("Hoja1",)
        if hasattr(self, "sheet_combo"):
            self.sheet_combo.configure(values=self._sheet_values)

        self._excel_rows_cache = []
        self._validation_rows_cache = []
        self._validation_tree_cache = []
        self._execution_cp_values = ()
        self._execution_series_values = ()
        self._execution_series_by_cp = {}
        self._selected_cp_filters = []
        self._selected_serie_filters = []
        self._last_summary = None
        self._operation_total = 0

        self._refresh_execution_filter_widgets()
        self._reset_initial_state()

    # ------------------------------------------------------------------ Config / execution
    def _build_config(self, *, simulation_only: bool) -> DossierConfig:
        pdf_sources = [
            {
                "document_name": "Descriptivo de pintura",
                "source_pdf_path": self._read_input(self.descriptivo_var),
                "target_folder": "5 Procedimiento de fabricación",
                "final_name_pattern": "5.2 Descriptivo de pintura.pdf",
                "mandatory": True,
            },
            {
                "document_name": "Comunicado técnico",
                "source_pdf_path": self._read_input(self.comunicado_var),
                "target_folder": "7 Ensayos",
                "final_name_pattern": "7.1 Comunicado técnico - Resultados de adherencia.pdf",
                "mandatory": True,
            },
            {
                "document_name": "Informe laboratorio",
                "source_pdf_path": self._read_input(self.informe_var),
                "target_folder": "7 Ensayos",
                "final_name_pattern": "7.2 Informe de Ensayo Laboratorio - prueba adherencia.pdf",
                "mandatory": True,
            },
        ]
        return DossierConfig.from_mapping(
            {
                "root_path": self._read_input(self.root_path_var),
                "excel_path": self._read_input(self.excel_path_var),
                "sheet_name": self.sheet_var.get().strip(),
                "dossier_folder_name": "06_DOSSIER",
                "simulation_only": simulation_only,
                "replace_existing": self.replace_var.get(),
                "cp_filter": self._selected_cp_filters[0] if len(self._selected_cp_filters) == 1 else "",
                "serie_filter": self._selected_serie_filters[0] if len(self._selected_serie_filters) == 1 else "",
                "cp_filters": list(self._selected_cp_filters),
                "serie_filters": list(self._selected_serie_filters),
                "cp_synonyms": [self.cp_column_var.get().strip() or "CP"],
                "serie_synonyms": [self.serie_column_var.get().strip() or "Serie"],
                "pdf_sources": pdf_sources,
            },
            source_path=Path(self.excel_path_var.get()).with_suffix(".json"),
        )

    def validate_cp_and_series(self) -> None:
        self._start_validation_operation()

    def simulate_distribution(self) -> None:
        self._start_operation(simulation_only=True, action_label="Simulando distribución")

    def execute_real_distribution(self) -> None:
        if not messagebox.askyesno(
            "Fase 2",
            "Está a punto de copiar y reemplazar documentos en la ruta documental real. Se crearán respaldos si corresponde. ¿Desea continuar?",
        ):
            return
        self._start_operation(simulation_only=False, action_label="Ejecutando distribución real")

    def _start_operation(self, *, simulation_only: bool, action_label: str) -> None:
        if self._operation_running:
            messagebox.showinfo("Fase 2", f"Ya hay una {self._current_operation_label} en curso.")
            return

        try:
            config = self._build_config(simulation_only=simulation_only)
            self._cancel_requested.clear()
            self._operation_running = True
            self._current_operation_label = "simulación" if simulation_only else "ejecución real"
            active_button = self.simulate_button if simulation_only else self.real_button
            self._set_running_state(True, active_button=active_button)
            self._set_operation_progress(0, 0, action_label)
            self.update_idletasks()

            if hasattr(self.winfo_toplevel(), "footer_mode_var"):
                self.winfo_toplevel().footer_mode_var.set("Simulación" if simulation_only else "Real")

            def progress_update(current: int, total: int, message: str) -> None:
                if self._cancel_requested.is_set():
                    raise OperationCancelledError("Operación cancelada por el usuario.")
                self._operation_queue.put(("progress", (current, total, message)))

            def worker() -> None:
                try:
                    phase1_items = self._build_phase1_items(config) if self.use_phase1_var.get() else []
                    if self._cancel_requested.is_set():
                        self._operation_queue.put(("cancelled", None))
                        return
                    summary = self.service.run_distribution(
                        config,
                        confirm_real=not simulation_only,
                        phase1_items=phase1_items,
                        progress_callback=progress_update,
                    )
                    if self._cancel_requested.is_set():
                        self._operation_queue.put(("cancelled", None))
                    else:
                        self._operation_queue.put(("done", summary))
                except OperationCancelledError:
                    self._operation_queue.put(("cancelled", None))
                except Exception as exc:  # pragma: no cover - background thread
                    self._operation_queue.put(("error", exc))

            self._operation_worker = threading.Thread(target=worker, daemon=True)
            self._operation_worker.start()
        except Exception as exc:
            logger.exception("Phase 2 execution failed")
            messagebox.showerror("Fase 2", str(exc))
            self._set_running_state(False)
            self._operation_running = False

    def _start_validation_operation(self) -> None:
        if self._operation_running:
            messagebox.showinfo("Fase 2", f"Ya hay una {self._current_operation_label} en curso.")
            return

        try:
            config = self._build_config(simulation_only=True)
            self._cancel_requested.clear()
            self._operation_running = True
            self._current_operation_label = "validación"
            self._set_running_state(True, active_button=self.validate_button)
            self._set_operation_progress(0, 0, "Validando CP y Series")
            self.update_idletasks()

            if hasattr(self.winfo_toplevel(), "footer_mode_var"):
                self.winfo_toplevel().footer_mode_var.set("Simulación")

            def progress_update(current: int, total: int, message: str) -> None:
                if self._cancel_requested.is_set():
                    raise OperationCancelledError("Operación cancelada por el usuario.")
                self._operation_queue.put(("progress", (current, total, message)))

            def worker() -> None:
                try:
                    _workbook_info, rows = self.validator.load_rows(config)
                    validated_rows = self.validator.validate_paths(config, rows, progress_callback=progress_update)
                    validation_tree = self.validator.build_validation_tree(config, validated_rows)
                    if self._cancel_requested.is_set():
                        self._operation_queue.put(("cancelled", None))
                    else:
                        self._operation_queue.put(("validated", (validated_rows, validation_tree)))
                except OperationCancelledError:
                    self._operation_queue.put(("cancelled", None))
                except Exception as exc:  # pragma: no cover - background thread
                    self._operation_queue.put(("error", exc))

            self._operation_worker = threading.Thread(target=worker, daemon=True)
            self._operation_worker.start()
        except Exception as exc:
            logger.exception("Phase 2 validation failed")
            messagebox.showerror("Fase 2", str(exc))
            self._set_running_state(False)
            self._operation_running = False

    def _build_phase1_items(self, config: DossierConfig) -> list[SimpleNamespace]:
        folder_value = self._read_input(self.phase1_folder_var)
        folder = Path(folder_value) if folder_value else Path()
        try:
            _workbook_info, rows = self.validator.load_rows(config)
            rows = self.service._filter_rows(config, rows)
        except Exception:
            return []

        pdf_files = [path for path in folder.rglob("*.pdf") if path.is_file()] if folder.exists() else []
        pdf_by_series: dict[str, Path] = {}
        for pdf in pdf_files:
            normalized_name = normalize_for_match(pdf.name)
            for token in normalized_name.replace("-", " ").replace("_", " ").split():
                if token.isdigit() and len(token) >= 6:
                    pdf_by_series.setdefault(token, pdf)
        items: list[SimpleNamespace] = []
        for row in rows:
            best_match = pdf_by_series.get(normalize_for_match(row.serie))
            expected_name = f"Certificado de Trazabilidad de Ensayos de Pintura SN {row.serie}.pdf"
            pdf_path = best_match or (folder / expected_name)
            items.append(
                SimpleNamespace(
                    row_number=row.row_number,
                    series=row.serie,
                    pdf_filename=pdf_path.name,
                    pdf_path=str(pdf_path),
                )
            )
        return items

    # ------------------------------------------------------------------ Queue / summary
    def _set_running_state(self, running: bool, active_button: ttk.Button | None = None) -> None:
        action_buttons = [getattr(self, "validate_button", None), getattr(self, "simulate_button", None), getattr(self, "real_button", None)]
        cancel_button = getattr(self, "cancel_button", None)

        if running:
            self._active_action_button = active_button
            for widget in action_buttons:
                if widget is None:
                    continue
                self._default_button_labels.setdefault(widget, widget.cget("text"))
                widget.configure(state="disabled")
            if active_button is not None:
                active_button.configure(state="normal", text=f"{self._default_button_labels.get(active_button, active_button.cget('text'))}...")
            if cancel_button is not None:
                self._default_button_labels.setdefault(cancel_button, cancel_button.cget("text"))
                cancel_button.configure(state="normal")
            return

        for widget in action_buttons + [cancel_button]:
            if widget is None:
                continue
            widget.configure(state="normal")
            if widget in self._default_button_labels:
                widget.configure(text=self._default_button_labels[widget])
        self._active_action_button = None

    def _poll_operation_queue(self) -> None:
        try:
            while True:
                kind, payload = self._operation_queue.get_nowait()
                if kind == "progress":
                    current, total, message = payload
                    self._update_operation_progress(int(current), int(total), str(message))
                    continue
                if kind == "validated":
                    self._operation_running = False
                    self._set_running_state(False)
                    rows, tree_rows = payload
                    self._apply_validation_rows(rows, tree_rows)
                    final_total = max(self._operation_total or len(rows), 1)
                    self._set_operation_progress(final_total, final_total, "Validación completada")
                    self._show_validation_dialog(tree_rows)
                    continue
                if kind == "done":
                    self._operation_running = False
                    self._set_running_state(False)
                    self._apply_summary(payload)
                    final_total = max(payload.planned_actions or len(payload.items), 1)
                    self._set_operation_progress(final_total, final_total, "Simulación completada" if payload.execution_mode == DossierExecutionMode.SIMULATION else "Ejecución real completada")
                    self._show_distribution_dialog(payload)
                    if self.open_report_var.get():
                        if getattr(payload, 'simulation_root', '') and payload.execution_mode == DossierExecutionMode.SIMULATION:
                            simulation_root = Path(getattr(payload, 'simulation_root', ''))
                            if simulation_root.exists():
                                self.after(200, lambda: self._open_path(str(simulation_root), "Simulación"))
                            else:
                                logger.warning("La carpeta de simulación no existe: %s", simulation_root)
                        else:
                            self.after(200, self.open_report)
                    continue
                if kind == "cancelled":
                    self._operation_running = False
                    self._set_running_state(False)
                    self._stop_progress_animation()
                    self.progress_state_var.set(f"{self._current_operation_label.capitalize()} cancelada")
                    self.progress_percent_var.set("Cancelado")
                    self.processed_var.set("Cancelado")
                    continue
                if kind == "error":
                    self._operation_running = False
                    self._set_running_state(False)
                    self.progress_state_var.set(f"La {self._current_operation_label} falló")
                    logger.exception("Phase 2 execution failed: %s", payload)
                    messagebox.showerror("Fase 2", str(payload))
        except queue.Empty:
            pass
        self.after(QUEUE_POLL_MS, self._poll_operation_queue)

    def _apply_summary(self, summary) -> None:
        self._last_summary = summary
        visible_result_path = str(summary.report_path or getattr(summary, 'simulation_root', '') or self.report_path_var.get())
        self.report_path_var.set(visible_result_path)
        self.last_update_var.set(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        self._write_log(summary)

        rows = self._preview_rows_from_summary(summary)
        self.preview_records = rows
        self._refresh_preview_tree()
        self._update_counts_from_preview()
        self._refresh_chart()

        processed = len(rows) if rows else max(summary.total_rows, 0)
        self.processed_var.set(f"{processed} / {processed}")
        self.progress_percent_var.set("100%")
        self.progress["value"] = 100
        self.progress_state_var.set("Simulación completada" if summary.execution_mode == DossierExecutionMode.SIMULATION else "Ejecución real completada")

        if hasattr(self.winfo_toplevel(), "footer_mode_var"):
            self.winfo_toplevel().footer_mode_var.set("Simulación" if summary.execution_mode == DossierExecutionMode.SIMULATION else "Real")

    def _apply_validation_rows(self, rows: list[DossierRow], tree_rows: list[DossierValidationTreeRow]) -> None:
        self._validation_rows_cache = rows
        self._validation_tree_cache = tree_rows
        self.preview_records = [self._validation_row_to_record(row) for row in rows]
        self._refresh_preview_tree()
        self._update_counts_from_preview()
        self._refresh_chart()

        total_rows = len(rows)
        self.processed_var.set(f"{total_rows} / {total_rows}")
        self.progress_percent_var.set("100%")
        self.progress["value"] = 100
        self.correct_var.set(str(sum(1 for row in rows if row.status == DossierStatus.VALID)))
        self.warning_var.set(str(sum(1 for row in rows if row.status == DossierStatus.WARNING)))
        self.error_var.set(str(sum(1 for row in rows if row.status == DossierStatus.ERROR)))
        self.skipped_var.set(str(sum(1 for row in rows if row.status == DossierStatus.SKIPPED)))
        self.last_update_var.set(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))

    def _set_operation_progress(self, current: int, total: int, message: str) -> None:
        total = max(int(total), 0)
        current = max(0, min(int(current), total if total else int(current)))
        self._operation_total = total
        if total <= 0:
            self.progress.configure(mode="indeterminate", maximum=100)
            self.progress.start(PROGRESS_ANIMATION_MS)
            self.progress["value"] = 0
            self.progress_percent_var.set("...")
            self.processed_var.set("Procesando...")
        else:
            self._stop_progress_animation()
            self.progress.configure(mode="determinate", maximum=100)
            percent = min(100, round((current / total) * 100))
            self.progress["value"] = percent
            self.progress_percent_var.set(f"{percent}%")
            self.processed_var.set(f"{current} / {total}")
        self.progress_state_var.set(message)

    def _update_operation_progress(self, current: int, total: int, message: str) -> None:
        total = max(int(total), 0)
        current = max(0, min(int(current), total if total else int(current)))
        self._operation_total = total
        if total <= 0:
            self._set_operation_progress(current, total, message)
            return
        self._stop_progress_animation()
        self.progress.configure(mode="determinate", maximum=100)
        percent = min(100, round((current / total) * 100))
        self.progress["value"] = percent
        self.progress_state_var.set(message)
        self.progress_percent_var.set(f"{percent}%")
        self.processed_var.set(f"{current} / {total}")

    def _show_validation_dialog(self, tree_rows: list[DossierValidationTreeRow]) -> None:
        nodes = self._build_validation_tree_nodes(tree_rows)
        summary_rows = self.preview_records or [self._validation_row_to_record(row) for row in self._validation_rows_cache]
        owner = self.winfo_toplevel()
        if hasattr(owner, '_bring_to_front'):
            owner._bring_to_front()
        DossierResultDialog(
            self.winfo_toplevel(),
            title="Resultado de validación CP y Series",
            subtitle="Vista detallada de candidatos CP, 06_DOSSIER, Planos, series y carpetas destino.",
            tree_nodes=nodes,
            summary_rows=summary_rows,
            summary_title="Resumen de validación",
        )

    def _show_distribution_dialog(self, summary) -> None:
        nodes = self._build_distribution_tree_nodes(summary)
        summary_rows = self._preview_rows_from_summary(summary)
        title = "Simulación de distribución" if summary.execution_mode == DossierExecutionMode.SIMULATION else "Resultado de distribución real"
        subtitle = "Vista detallada de origen, destino, nombre final, acción y respaldo planificado o ejecutado."
        DossierResultDialog(
            self.winfo_toplevel(),
            title=title,
            subtitle=subtitle,
            tree_nodes=nodes,
            summary_rows=summary_rows,
            summary_title="Resumen de distribución",
        )

    def _build_validation_tree_nodes(self, tree_rows: list[DossierValidationTreeRow]) -> list[dict[str, object]]:
        nodes: list[dict[str, object]] = []
        for row in tree_rows:
            row_status_label = STATUS_LABELS.get(row.status, row.status.value.title()) if isinstance(row.status, DossierStatus) else str(row.status)
            candidate_count = len(row.candidates)
            row_detail = f"Estado: {row_status_label} | {row.observation or 'Sin observaciones'} | Carpetas CP revisadas: {candidate_count}"
            row_children: list[dict[str, object]] = []
            if row.selected_cp_folder:
                row_children.append(
                    {
                        "text": f"Carpeta CP válida: {Path(row.selected_cp_folder).name}",
                        "status": "correct",
                        "detail": "Carpeta seleccionada para distribución.",
                        "children": [],
                    }
                )
            elif row.match_state == "ambiguous":
                row_children.append(
                    {
                        "text": "Carpeta CP no seleccionada",
                        "status": "warning",
                        "detail": f"Coincidencia ambigua en {row.cp}.",
                        "children": [],
                    }
                )
            elif row.match_state == "missing":
                row_children.append(
                    {
                        "text": "Carpeta CP no encontrada",
                        "status": "error",
                        "detail": f"No se encontró una carpeta candidata para {row.cp}.",
                        "children": [],
                    }
                )

            if not row.candidates:
                row_children.append(
                    {
                        "text": "Sin candidatos CP",
                        "status": "error",
                        "detail": "No se encontraron carpetas CP candidatas en la ruta raíz.",
                        "children": [],
                    }
                )

            for candidate in row.candidates:
                candidate_status = "correct" if candidate.valid_for_distribution else ("warning" if candidate.dossier_exists else "error")
                series_location_nodes = [
                    {"text": Path(location).name, "status": "correct", "detail": location, "children": []}
                    for location in getattr(candidate, "series_locations", [])[:20]
                ]
                candidate_children = [
                    {"text": "06_DOSSIER", "status": "correct" if candidate.dossier_exists else "error", "detail": candidate.dossier_folder or ("Sí" if candidate.dossier_exists else "No"), "children": []},
                    {"text": "Planos", "status": "correct" if candidate.planos_exists else "error", "detail": candidate.planos_folder or ("Sí" if candidate.planos_exists else "No"), "children": []},
                    {"text": "Serie encontrada", "status": "correct" if candidate.series_found else "error", "detail": "Sí" if candidate.series_found else "No", "children": series_location_nodes},
                    {"text": "Carpeta 5", "status": "correct" if candidate.folder_5_exists else "warning", "detail": candidate.folder_5 or "No", "children": []},
                    {"text": "Carpeta 6", "status": "correct" if candidate.folder_6_exists else "warning", "detail": candidate.folder_6 or "No", "children": []},
                    {"text": "Carpeta 7", "status": "correct" if candidate.folder_7_exists else "warning", "detail": candidate.folder_7 or "No", "children": []},
                ]
                candidate_label = Path(candidate.cp_folder).name if candidate.cp_folder else candidate.cp_folder
                if candidate.selected:
                    candidate_label = f"{candidate_label} (seleccionada)"
                row_children.append(
                    {
                        "text": candidate_label,
                        "status": candidate_status,
                        "detail": candidate.reason,
                        "children": candidate_children,
                    }
                )

            nodes.append(
                {
                    "text": f"Fila {row.row_number} · CP {row.cp} · Serie {row.serie}",
                    "status": self._status_tag(row.status),
                    "detail": row_detail,
                    "children": row_children,
                }
            )
        return nodes

    def _build_distribution_tree_nodes(self, summary) -> list[dict[str, object]]:
        grouped: dict[int, list[dict[str, object]]] = defaultdict(list)
        row_meta: dict[int, dict[str, str]] = {}
        for item in summary.items:
            row_number = int(getattr(item, "row_number", 0) or 0)
            if row_number <= 0:
                continue
            row_meta.setdefault(row_number, {"cp": getattr(item, "cp", ""), "serie": getattr(item, "serie", "")})
            action_label = getattr(getattr(item, "action_type", None), "value", getattr(item, "action_type", ""))
            status = getattr(item, "status", DossierStatus.SKIPPED)
            status_label = STATUS_LABELS.get(status, status.value.title()) if isinstance(status, DossierStatus) else str(status)
            detail = (
                f"Origen: {getattr(item, 'source_pdf_path', '') or 'No aplica'} · "
                f"Destino: {getattr(item, 'planned_path', '') or 'No aplica'} · "
                f"Respaldo: {getattr(item, 'backup_path', '') or 'No aplica'} · "
                f"Acción: {action_label or 'No aplica'} · Estado: {status_label}"
            )
            grouped[row_number].append(
                {
                    "text": getattr(item, "rule_name", "Acción") or "Acción",
                    "status": "correct" if status in {DossierStatus.PLANNED, DossierStatus.COPIED, DossierStatus.REPLACED} else ("warning" if status == DossierStatus.SKIPPED else "error"),
                    "detail": detail + (f" · {getattr(item, 'observation', '') or getattr(item, 'skipped_reason', '')}" if getattr(item, 'observation', '') or getattr(item, 'skipped_reason', '') else ""),
                    "children": [],
                }
            )

        nodes: list[dict[str, object]] = []
        for row_number in sorted(grouped):
            meta = row_meta.get(row_number, {"cp": "", "serie": ""})
            row_status_tag = self._aggregate_status_tag([child.get("status") for child in grouped[row_number]])
            nodes.append(
                {
                    "text": f"Fila {row_number} · CP {meta['cp']} · Serie {meta['serie']}",
                    "status": row_status_tag,
                    "detail": f"Acciones planificadas o ejecutadas: {len(grouped[row_number])}",
                    "children": grouped[row_number],
                }
            )
        return nodes

    @staticmethod
    def _aggregate_status_tag(statuses: list[object]) -> str:
        priority = {"error": 3, "warning": 2, "correct": 1, "skipped": 0}
        best = "skipped"
        best_score = -1
        for status in statuses:
            tag = status if isinstance(status, str) else "skipped"
            score = priority.get(tag, 0)
            if score > best_score:
                best = tag
                best_score = score
        return best

    def _validation_row_to_record(self, row: DossierRow) -> dict[str, object]:
        if row.cp_folder and row.series_in_planos:
            carpeta_cp = row.cp_folder
        elif row.observation == "CP no encontrada en la ruta documental":
            carpeta_cp = "No encontrada"
        elif row.observation == "Serie no encontrada en 06_DOSSIER para esta CP":
            carpeta_cp = "CP encontrada, serie no localizada"
        elif row.observation == "Existe 06_DOSSIER, pero no se encontró carpeta Planos":
            carpeta_cp = "CP encontrada, pero no existe 06_DOSSIER"
        else:
            carpeta_cp = row.cp_folder or "No encontrada"

        if row.status == DossierStatus.VALID and row.series_in_planos:
            estado = DossierStatus.VALID if all((row.folder_5, row.folder_6, row.folder_7)) else DossierStatus.WARNING
            if estado == DossierStatus.VALID:
                observacion = "Lista para simular distribución"
            else:
                faltantes = []
                if not row.folder_5:
                    faltantes.append("carpeta 5")
                if not row.folder_6:
                    faltantes.append("carpeta 6")
                if not row.folder_7:
                    faltantes.append("carpeta 7")
                observacion = f"Faltan carpetas destino: {', '.join(faltantes)}"
        elif row.observation == "CP no encontrada en la ruta documental":
            estado = DossierStatus.SKIPPED
            observacion = row.observation
        elif row.observation == "Serie no encontrada en 06_DOSSIER para esta CP":
            estado = DossierStatus.SKIPPED
            observacion = row.observation
        elif row.observation == "Existe 06_DOSSIER, pero no se encontró carpeta Planos":
            estado = DossierStatus.SKIPPED
            observacion = row.observation
        else:
            estado = row.status if row.status in {DossierStatus.WARNING, DossierStatus.ERROR} else DossierStatus.SKIPPED
            observacion = row.observation or "No se pudo validar la fila"

        carpeta_5 = "Sí" if row.folder_5 else ("-" if estado == DossierStatus.SKIPPED else "No")
        carpeta_6 = "Sí" if row.folder_6 else ("-" if estado == DossierStatus.SKIPPED else "No")
        carpeta_7 = "Sí" if row.folder_7 else ("-" if estado == DossierStatus.SKIPPED else "No")
        accion = "4 documentos" if estado == DossierStatus.VALID else ("0 documentos" if estado == DossierStatus.SKIPPED else (f"{sum(bool(x) for x in (row.folder_5, row.folder_6, row.folder_7))} documentos"))

        return {
            "fila_excel": row.row_number,
            "cp": row.cp,
            "serie": row.serie,
            "carpeta_cp": carpeta_cp,
            "serie_planos": "Sí" if row.series_in_planos else "No",
            "carpeta_5": carpeta_5,
            "carpeta_6": carpeta_6,
            "carpeta_7": carpeta_7,
            "accion": accion,
            "estado": estado,
            "observacion": observacion,
        }

    def _start_progress_animation(self) -> None:
        try:
            self.progress.configure(mode="indeterminate")
            self.progress.start(PROGRESS_ANIMATION_MS)
        except Exception:
            pass

    def _stop_progress_animation(self) -> None:
        try:
            self.progress.stop()
            self.progress.configure(mode="determinate")
        except Exception:
            pass

    def _preview_rows_from_summary(self, summary) -> list[dict[str, object]]:
        grouped: dict[int, dict[str, object]] = {}
        for item in summary.items:
            row_number = int(getattr(item, "row_number", 0) or 0)
            if row_number <= 0:
                continue
            bucket = grouped.setdefault(
                row_number,
                {
                    "fila_excel": row_number,
                    "cp": getattr(item, "cp", ""),
                    "serie": getattr(item, "serie", ""),
                    "carpeta_cp": "No encontrada",
                    "serie_planos": "No",
                    "carpeta_5": "-",
                    "carpeta_6": "-",
                    "carpeta_7": "-",
                    "accion": 0,
                    "estado": DossierStatus.SKIPPED,
                    "observacion": "",
                    "_statuses": [],
                    "_messages": [],
                    "_success_messages": [],
                    "_warning_messages": [],
                    "_error_messages": [],
                    "_skipped_messages": [],
                },
            )

            status = getattr(item, "status", DossierStatus.SKIPPED)
            bucket["_statuses"].append(status)

            message = getattr(item, "observation", "") or getattr(item, "skipped_reason", "") or ""
            if message:
                bucket["_messages"].append(message)
                if status in {DossierStatus.COPIED, DossierStatus.REPLACED, DossierStatus.PLANNED, DossierStatus.VALID, DossierStatus.SIMULATED}:
                    bucket["_success_messages"].append(message)
                elif status == DossierStatus.WARNING:
                    bucket["_warning_messages"].append(message)
                elif status in {DossierStatus.ERROR, DossierStatus.BLOCKED}:
                    bucket["_error_messages"].append(message)
                else:
                    bucket["_skipped_messages"].append(message)

            if getattr(item, "planned_path", ""):
                planned = Path(item.planned_path)
                if len(planned.parents) >= 3:
                    bucket["carpeta_cp"] = str(planned.parents[2])
                bucket["serie_planos"] = "Sí"
                target_folder = normalize_for_match(getattr(item, "target_folder", ""))
                if target_folder.startswith("5"):
                    bucket["carpeta_5"] = "Sí"
                elif target_folder.startswith("6"):
                    bucket["carpeta_6"] = "Sí"
                elif target_folder.startswith("7"):
                    bucket["carpeta_7"] = "Sí"
                if getattr(item, "action_type", None) and getattr(item, "action_type").name in {"COPY", "REPLACE", "PLANNED"}:
                    bucket["accion"] += 1
            elif getattr(item, "rule_name", "") == "phase1-routing" and getattr(item, "status", None) == DossierStatus.SKIPPED:
                bucket["carpeta_6"] = bucket["carpeta_6"] if bucket["carpeta_6"] != "-" else "No"

        result: list[dict[str, object]] = []
        for row_number in sorted(grouped):
            bucket = grouped[row_number]
            statuses = bucket.pop("_statuses", [])
            success_messages = bucket.pop("_success_messages", [])
            warning_messages = bucket.pop("_warning_messages", [])
            error_messages = bucket.pop("_error_messages", [])
            skipped_messages = bucket.pop("_skipped_messages", [])
            all_messages = bucket.pop("_messages", [])

            action_count = bucket["accion"]
            has_error = any(status in {DossierStatus.ERROR, DossierStatus.BLOCKED} for status in statuses)
            has_warning = any(status == DossierStatus.WARNING for status in statuses)
            has_skipped = any(status == DossierStatus.SKIPPED for status in statuses)
            success_statuses = [status for status in statuses if status in {DossierStatus.COPIED, DossierStatus.REPLACED, DossierStatus.PLANNED, DossierStatus.VALID, DossierStatus.SIMULATED}]
            has_success = bool(success_statuses)

            if has_error:
                bucket["estado"] = DossierStatus.ERROR
                bucket["observacion"] = error_messages[-1] if error_messages else (all_messages[-1] if all_messages else "La fila presentó errores.")
            elif has_success and (has_warning or has_skipped):
                bucket["estado"] = DossierStatus.WARNING
                success_total = len(success_statuses)
                skipped_total = sum(1 for status in statuses if status in {DossierStatus.WARNING, DossierStatus.SKIPPED})
                detail_message = warning_messages[-1] if warning_messages else (skipped_messages[-1] if skipped_messages else "Revisá el detalle.")
                bucket["observacion"] = f"Se agregaron correctamente {success_total} documento(s). {skipped_total} quedaron pendientes u omitidos. {detail_message}"
            elif has_success:
                bucket["estado"] = self._best_success_status(success_statuses)
                bucket["observacion"] = success_messages[-1] if success_messages else "Agregado correctamente."
            elif has_warning:
                bucket["estado"] = DossierStatus.WARNING
                bucket["observacion"] = warning_messages[-1] if warning_messages else "Revisá el detalle de la fila."
            else:
                bucket["estado"] = DossierStatus.SKIPPED
                bucket["observacion"] = skipped_messages[-1] if skipped_messages else (all_messages[-1] if all_messages else "No se agregó ningún documento.")

            if bucket["estado"] == DossierStatus.VALID and action_count == 0:
                action_count = 4 if self.use_phase1_var.get() else 3
            bucket["accion"] = f"{action_count} documentos" if action_count else "0 documentos"
            if has_success:
                bucket["serie_planos"] = "Sí"
            result.append(bucket)
        return result

    @staticmethod
    def _best_success_status(statuses: list[DossierStatus]) -> DossierStatus:
        if DossierStatus.REPLACED in statuses:
            return DossierStatus.REPLACED
        if DossierStatus.COPIED in statuses:
            return DossierStatus.COPIED
        if DossierStatus.PLANNED in statuses:
            return DossierStatus.PLANNED
        if DossierStatus.SIMULATED in statuses:
            return DossierStatus.SIMULATED
        return DossierStatus.VALID

    def _update_counts_from_preview(self) -> None:
        counts = Counter(record["estado"] for record in self.preview_records)
        self.chart_counts = Counter(
            {
                DossierStatus.VALID: counts.get(DossierStatus.VALID, 0) + counts.get(DossierStatus.PLANNED, 0) + counts.get(DossierStatus.COPIED, 0) + counts.get(DossierStatus.REPLACED, 0),
                DossierStatus.WARNING: counts.get(DossierStatus.WARNING, 0),
                DossierStatus.ERROR: counts.get(DossierStatus.ERROR, 0) + counts.get(DossierStatus.BLOCKED, 0),
                DossierStatus.SKIPPED: counts.get(DossierStatus.SKIPPED, 0),
            }
        )
        self.correct_var.set(str(self.chart_counts.get(DossierStatus.VALID, 0)))
        self.warning_var.set(str(self.chart_counts.get(DossierStatus.WARNING, 0)))
        self.error_var.set(str(self.chart_counts.get(DossierStatus.ERROR, 0)))
        self.skipped_var.set(str(self.chart_counts.get(DossierStatus.SKIPPED, 0)))

    def _write_log(self, summary) -> None:
        log_root = Path(r"C:\Automatizacion\Logs")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_path = log_root / f"log_distribucion_{timestamp}.txt"
        try:
            log_root.mkdir(parents=True, exist_ok=True)
            with log_path.open("w", encoding="utf-8") as handle:
                handle.write("AutomatizaciónDocumental - Fase 2\n")
                handle.write(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                handle.write(f"Modo: {summary.execution_mode.value}\n")
                handle.write(f"Total filas: {summary.total_rows}\n")
                handle.write(f"Correctos: {summary.valid_rows}\n")
                handle.write(f"Advertencias: {summary.warnings}\n")
                handle.write(f"Errores: {summary.errors}\n")
                handle.write(f"Bloqueados: {summary.blocked}\n")
                handle.write(f"Acciones planificadas: {summary.planned_actions}\n")
                handle.write("\nDetalle:\n")
                for item in summary.items:
                    handle.write(f"- {item.row_number} | {item.cp} | {item.serie} | {item.status.value} | {item.observation or item.skipped_reason}\n")
            self.log_path_var.set(str(log_path))
        except Exception as exc:  # pragma: no cover - filesystem issues
            logger.exception("Unable to write log file: %s", exc)

    def _clear_preview_filter(self) -> None:
        self.preview_filter_var.set("")
        self._refresh_preview_tree()

    def _refresh_preview_tree(self) -> None:
        if not hasattr(self, "preview_tree"):
            return
        self.preview_tree.delete(*self.preview_tree.get_children())
        filter_text = normalize_for_match(self.preview_filter_var.get())
        for record in self.preview_records:
            haystack = normalize_for_match(" ".join(str(record.get(key, "")) for key in record.keys()))
            if filter_text and filter_text not in haystack:
                continue
            status = record["estado"]
            tag = STATUS_TAGS.get(status, "skipped")
            self.preview_tree.insert(
                "",
                "end",
                values=(
                    record["fila_excel"],
                    record["cp"],
                    record["serie"],
                    record["carpeta_cp"],
                    record["serie_planos"],
                    record["carpeta_5"],
                    record["carpeta_6"],
                    record["carpeta_7"],
                    record["accion"],
                    STATUS_LABELS.get(status, str(status)),
                    record["observacion"],
                ),
                tags=(tag,),
            )

    # ------------------------------------------------------------------ Open / cancel
    def cancel_operation(self) -> None:
        if self._operation_running:
            messagebox.showinfo("Fase 2", "La cancelación no está disponible durante la ejecución en curso.")
            return
        messagebox.showinfo("Fase 2", "No hay un proceso activo.")

    def open_report(self) -> None:
        self._open_path(self.report_path_var.get(), "Reporte")

    def open_log(self) -> None:
        self._open_path(self.log_path_var.get(), "Log")

    def _open_path(self, path_value: str, label: str) -> None:
        path = Path(path_value)
        if not path.exists():
            messagebox.showwarning(label, f"La ruta no existe:\n{path}")
            return
        try:
            if hasattr(os, "startfile"):
                os.startfile(path)  # type: ignore[attr-defined]
                return
            messagebox.showinfo(label, str(path))
        except Exception as exc:  # pragma: no cover - platform specific
            logger.exception("No se pudo abrir la ruta %s", path)
            messagebox.showerror(label, f"No se pudo abrir la ruta:\n{path}\n\n{exc}")


class DossierResultDialog(tk.Toplevel):
    def __init__(
        self,
        parent: tk.Widget,
        *,
        title: str,
        subtitle: str,
        tree_nodes: list[dict[str, object]],
        summary_rows: list[dict[str, object]] | None = None,
        summary_title: str = "Resumen",
    ) -> None:
        super().__init__(parent)
        self.withdraw()
        self.title(title)
        self.configure(bg=PALETTE["background"])
        self.geometry("1280x780")
        self.minsize(1120, 680)
        self.resizable(True, True)
        self.protocol("WM_DELETE_WINDOW", self._close_dialog)
        self.bind("<Escape>", lambda _event: self._close_dialog())

        container = ttk.Frame(self, style="App.TFrame", padding=16)
        container.pack(fill="both", expand=True)
        container.rowconfigure(1, weight=3)
        container.rowconfigure(3, weight=2)
        container.columnconfigure(0, weight=1)

        header = ttk.Frame(container, style="CardBody.TFrame")
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=1)
        tk.Label(header, text=title, bg=PALETTE["surface"], fg=PALETTE["accent"], font=("Segoe UI Semibold", 13)).pack(anchor="w", padx=12, pady=(10, 2))
        tk.Label(header, text=subtitle, bg=PALETTE["surface"], fg=PALETTE["muted"], font=("Segoe UI", 9), wraplength=1160, justify="left").pack(anchor="w", padx=12, pady=(0, 10))

        tree_frame = ttk.Frame(container, style="CardBody.TFrame")
        tree_frame.grid(row=1, column=0, sticky="nsew", pady=(12, 10))
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        tree_columns = ("estado", "detalle")
        self.tree = ttk.Treeview(tree_frame, columns=tree_columns, show="tree headings", style="App.Treeview")
        self.tree.heading("#0", text="Elemento")
        self.tree.heading("estado", text="Estado")
        self.tree.heading("detalle", text="Detalle")
        self.tree.column("#0", width=360, anchor="w", stretch=True)
        self.tree.column("estado", width=120, anchor="center", stretch=False)
        self.tree.column("detalle", width=650, anchor="w", stretch=True)
        self.tree.grid(row=0, column=0, sticky="nsew")

        y_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        self.tree.tag_configure("correct", background=PALETTE["success"])
        self.tree.tag_configure("warning", background=PALETTE["warning"])
        self.tree.tag_configure("error", background=PALETTE["danger"])
        self.tree.tag_configure("skipped", background=PALETTE["accent_light"])

        self._insert_nodes("", tree_nodes)

        summary_frame = ttk.Frame(container, style="CardBody.TFrame")
        summary_frame.grid(row=3, column=0, sticky="nsew")
        summary_frame.rowconfigure(1, weight=1)
        summary_frame.columnconfigure(0, weight=1)

        tk.Label(summary_frame, text=summary_title, bg=PALETTE["surface"], fg=PALETTE["accent"], font=("Segoe UI Semibold", 11)).grid(row=0, column=0, sticky="w", padx=2, pady=(0, 6))
        if summary_rows:
            summary_tree = ttk.Treeview(
                summary_frame,
                columns=("fila_excel", "cp", "serie", "carpeta_cp", "serie_planos", "carpeta_5", "carpeta_6", "carpeta_7", "accion", "estado", "observacion"),
                show="headings",
                style="App.Treeview",
                selectmode="browse",
            )
            headings = [
                ("fila_excel", "Fila", 60),
                ("cp", "CP", 120),
                ("serie", "Serie", 120),
                ("carpeta_cp", "Carpeta CP", 180),
                ("serie_planos", "Serie encontrada", 110),
                ("carpeta_5", "5", 55),
                ("carpeta_6", "6", 55),
                ("carpeta_7", "7", 55),
                ("accion", "Acción", 110),
                ("estado", "Estado", 110),
                ("observacion", "Observación", 240),
            ]
            for column, heading, width in headings:
                summary_tree.heading(column, text=heading)
                summary_tree.column(column, width=width, anchor="w", stretch=column == "observacion")
            summary_tree.grid(row=1, column=0, sticky="nsew")

            summary_scroll = ttk.Scrollbar(summary_frame, orient="vertical", command=summary_tree.yview)
            summary_scroll.grid(row=1, column=1, sticky="ns")
            summary_tree.configure(yscrollcommand=summary_scroll.set)

            for record in summary_rows:
                estado = record.get("estado")
                status_tag = self._status_tag(estado)
                summary_tree.insert(
                    "",
                    "end",
                    values=(
                        record.get("fila_excel", ""),
                        record.get("cp", ""),
                        record.get("serie", ""),
                        record.get("carpeta_cp", ""),
                        record.get("serie_planos", ""),
                        record.get("carpeta_5", ""),
                        record.get("carpeta_6", ""),
                        record.get("carpeta_7", ""),
                        record.get("accion", ""),
                        self._display_status(estado),
                        record.get("observacion", ""),
                    ),
                    tags=(status_tag,),
                )
            summary_tree.tag_configure("correct", background=PALETTE["success"])
            summary_tree.tag_configure("warning", background=PALETTE["warning"])
            summary_tree.tag_configure("error", background=PALETTE["danger"])
            summary_tree.tag_configure("skipped", background=PALETTE["accent_light"])
        else:
            tk.Label(summary_frame, text="No hay resumen para mostrar.", bg=PALETTE["surface"], fg=PALETTE["muted"], font=("Segoe UI", 9)).grid(row=1, column=0, sticky="w")

        actions = ttk.Frame(container, style="CardBody.TFrame")
        actions.grid(row=4, column=0, sticky="e", pady=(12, 0))
        ttk.Button(actions, text="Cerrar", style="Ghost.TButton", command=self._close_dialog).pack(side="right")

        self.after(0, self._show_centered)

    def _insert_nodes(self, parent_id: str, nodes: list[dict[str, object]]) -> None:
        for node in nodes:
            status = self._status_tag(node.get("status"))
            item_id = self.tree.insert(
                parent_id,
                "end",
                text=str(node.get("text", "")),
                values=(self._display_status(node.get("status")), str(node.get("detail", ""))),
                tags=(status,),
                open=True,
            )
            children = node.get("children", [])
            if isinstance(children, list):
                self._insert_nodes(item_id, children)

    @staticmethod
    def _status_tag(status: object) -> str:
        if isinstance(status, DossierStatus):
            return {
                DossierStatus.VALID: "correct",
                DossierStatus.PLANNED: "correct",
                DossierStatus.COPIED: "correct",
                DossierStatus.REPLACED: "correct",
                DossierStatus.WARNING: "warning",
                DossierStatus.ERROR: "error",
                DossierStatus.BLOCKED: "error",
                DossierStatus.SKIPPED: "skipped",
                DossierStatus.SIMULATED: "correct",
            }.get(status, "skipped")
        if isinstance(status, str):
            return status if status in {"correct", "warning", "error", "skipped"} else "skipped"
        return "skipped"

    @staticmethod
    def _display_status(status: object) -> str:
        if isinstance(status, DossierStatus):
            return STATUS_LABELS.get(status, status.value.title())
        return {
            "correct": "Correcto",
            "warning": "Advertencia",
            "error": "Error",
            "skipped": "Omitido",
        }.get(str(status), str(status))

    def _show_centered(self) -> None:
        self.update_idletasks()
        parent = self.master.winfo_toplevel() if self.master is not None else None
        width = max(self.winfo_width(), 1120)
        height = max(self.winfo_height(), 680)

        if parent is not None and parent.winfo_exists():
            try:
                parent.update_idletasks()
                base_x = parent.winfo_rootx()
                base_y = parent.winfo_rooty()
                base_width = max(parent.winfo_width(), width)
                base_height = max(parent.winfo_height(), height)
                x = max(base_x + (base_width - width) // 2, 0)
                y = max(base_y + (base_height - height) // 2, 0)
            except Exception:
                x = max((self.winfo_screenwidth() - width) // 2, 0)
                y = max((self.winfo_screenheight() - height) // 2, 0)
        else:
            x = max((self.winfo_screenwidth() - width) // 2, 0)
            y = max((self.winfo_screenheight() - height) // 2, 0)

        self.geometry(f"{width}x{height}+{x}+{y}")
        self.deiconify()
        self.lift()
        try:
            self.attributes("-topmost", True)
            self.after(250, lambda: self.attributes("-topmost", False))
        except Exception:
            pass
        try:
            self.focus_force()
        except Exception:
            pass

    def _close_dialog(self) -> None:
        try:
            self.destroy()
        except Exception:
            pass

