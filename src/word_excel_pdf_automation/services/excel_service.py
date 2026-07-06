from __future__ import annotations

import logging
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from ..config import SERIAL_SYNONYMS
from ..models import ColumnInfo, SheetInfo, SeriesRow, ValidationStatus, WorkbookInfo
from ..utils.text import normalize_for_match, normalize_series, similarity


logger = logging.getLogger(__name__)


class ExcelService:
    def inspect_workbook(self, excel_path: Path) -> WorkbookInfo:
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        sheets: list[SheetInfo] = []
        for worksheet in workbook.worksheets:
            header_row = self._detect_header_row(worksheet.iter_rows(values_only=True))
            columns = self._extract_columns(worksheet, header_row)
            suggested = self._suggest_serial_column(columns)
            sheets.append(
                SheetInfo(
                    name=worksheet.title,
                    header_row=header_row,
                    columns=columns,
                    suggested_serial_column=suggested,
                )
            )
        workbook.close()
        return WorkbookInfo(path=excel_path, sheets=sheets)

    def load_series_rows(self, excel_path: Path, sheet_name: str, serial_column: str) -> list[SeriesRow]:
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        worksheet = workbook[sheet_name]
        header_row = self._detect_header_row(worksheet.iter_rows(values_only=True))
        columns = self._extract_columns(worksheet, header_row)
        column_index = self._resolve_column_index(columns, serial_column)

        seen = Counter()
        rows: list[SeriesRow] = []
        for row_number in range(header_row + 1, worksheet.max_row + 1):
            raw_value = worksheet.cell(row=row_number, column=column_index).value
            series = normalize_series(raw_value)
            normalized = normalize_for_match(series)

            if not series:
                status = ValidationStatus.EMPTY
                selected = False
                observation = "Serie vacía"
            else:
                seen[normalized] += 1
                if seen[normalized] > 1:
                    status = ValidationStatus.DUPLICATE
                    selected = False
                    observation = "Serie duplicada"
                else:
                    status = ValidationStatus.VALID
                    selected = True
                    observation = "Lista"

            rows.append(
                SeriesRow(
                    row_number=row_number,
                    series=series,
                    normalized_series=normalized,
                    selected=selected,
                    status=status,
                    observation=observation,
                    source_sheet=sheet_name,
                )
            )

        workbook.close()
        return rows

    @staticmethod
    def _detect_header_row(rows: Iterable[tuple]) -> int:
        best_row = 1
        best_score = float("-inf")
        for index, row in enumerate(rows, start=1):
            if index > 25:
                break
            values = [normalize_series(value) for value in row]
            non_empty = [value for value in values if value]
            if not non_empty:
                continue
            score = len(non_empty)
            normalized_row = [normalize_for_match(value) for value in non_empty]
            for candidate in normalized_row:
                if candidate in {normalize_for_match(s) for s in SERIAL_SYNONYMS}:
                    score += 25
                else:
                    score += max((similarity(candidate, synonym) for synonym in SERIAL_SYNONYMS), default=0.0) * 5
            if any(value.isdigit() for value in non_empty):
                score -= 2
            if score > best_score:
                best_score = score
                best_row = index
        return best_row

    @staticmethod
    def _extract_columns(worksheet, header_row: int) -> list[ColumnInfo]:
        columns: list[ColumnInfo] = []
        header_values = next(
            worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True),
            (),
        )
        for index, value in enumerate(header_values, start=1):
            name = normalize_series(value) or f"Column {index}"
            columns.append(ColumnInfo(index=index, name=name))
        return columns

    @staticmethod
    def _suggest_serial_column(columns: list[ColumnInfo]) -> str | None:
        best_name: str | None = None
        best_score = -1.0
        for column in columns:
            candidate = column.name
            score = 0.0
            normalized = normalize_for_match(candidate)
            for synonym in SERIAL_SYNONYMS:
                if normalized == normalize_for_match(synonym):
                    score += 100.0
                else:
                    score = max(score, similarity(candidate, synonym) * 100.0)
            if score > best_score:
                best_score = score
                best_name = candidate
        return best_name

    @staticmethod
    def _resolve_column_index(columns: list[ColumnInfo], serial_column: str) -> int:
        target = normalize_for_match(serial_column)
        for column in columns:
            if normalize_for_match(column.name) == target:
                return column.index
        raise ValueError(f"La columna de serie '{serial_column}' no se encontró en la hoja seleccionada.")
