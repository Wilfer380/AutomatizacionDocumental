from __future__ import annotations

from collections import Counter, defaultdict
import os
from dataclasses import dataclass, field
import re
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from ..dossier_models import DossierColumnInfo, DossierConfig, DossierRow, DossierStatus, DossierWorkbookInfo
from ..utils.text import normalize_for_match, normalize_series, similarity

_CP_IDENTIFIER_PATTERN = re.compile(r"\bcp[\s-]*(?P<code>\d+)\b")


class DossierWorkbookHeaderError(RuntimeError):
    pass


@dataclass(slots=True)
class DossierValidationCandidate:
    cp_folder: str
    dossier_folder: str
    planos_folder: str
    dossier_exists: bool
    planos_exists: bool
    series_found: bool
    folder_5_exists: bool
    folder_6_exists: bool
    folder_7_exists: bool
    selected: bool = False
    valid_for_distribution: bool = False
    reason: str = ""
    folder_5: str = ""
    folder_6: str = ""
    folder_7: str = ""
    series_locations: list[str] = field(default_factory=list)


@dataclass(slots=True)
class DossierValidationTreeRow:
    row_number: int
    cp: str
    serie: str
    selected_cp_folder: str
    match_state: str
    status: DossierStatus
    observation: str
    candidates: list[DossierValidationCandidate] = field(default_factory=list)


@dataclass(slots=True)
class DossierIndexedCandidate:
    cp_folder: Path
    cp_name_key: str
    cp_code: str
    dossier_folder: Path | None
    planos_folder: Path | None
    folder_5: Path | None
    folder_6: Path | None
    folder_7: Path | None
    series_keys: tuple[str, ...] = ()


@dataclass(slots=True)
class DossierRuntimeIndex:
    candidates: tuple[DossierIndexedCandidate, ...]
    cp_code_lookup: dict[str, tuple[DossierIndexedCandidate, ...]]


class DossierValidatorService:
    def find_cp_folders(self, root_path: Path, cp: str) -> list[Path]:
        if not root_path.exists():
            return []
        runtime_index = self._build_runtime_index(root_path)
        unique_matches: list[Path] = []
        seen_paths: set[str] = set()
        for candidate in self._match_cp_candidates(runtime_index, cp):
            path_key = str(candidate.cp_folder)
            if path_key in seen_paths:
                continue
            seen_paths.add(path_key)
            unique_matches.append(candidate.cp_folder)
        return unique_matches

    def find_dossier_folder(self, cp_folder: Path) -> Path | None:
        dossier_folders = self.find_dossier_folders(cp_folder)
        return dossier_folders[0] if dossier_folders else None

    def find_dossier_folders(self, cp_folder: Path) -> list[Path]:
        if not cp_folder.is_dir():
            return []
        matches: list[tuple[Path, float, int]] = []
        seen_paths: set[str] = set()
        for root, dirs, _files in self._safe_walk(cp_folder):
            root_score = self._score_directory_name(root.name, "06_DOSSIER")
            if root_score >= 0.45:
                path_key = str(root)
                if path_key not in seen_paths:
                    seen_paths.add(path_key)
                    matches.append((root, root_score + 0.2, len(root.parts)))
            elif self._looks_like_distribution_root(root, dirs):
                path_key = str(root)
                if path_key not in seen_paths:
                    seen_paths.add(path_key)
                    matches.append((root, 0.65, len(root.parts)))

            for directory_name in dirs:
                candidate = root / directory_name
                score = self._score_directory_name(directory_name, "06_DOSSIER")
                if score < 0.45:
                    continue
                path_key = str(candidate)
                if path_key in seen_paths:
                    continue
                seen_paths.add(path_key)
                matches.append((candidate, score, len(candidate.parts)))
        matches.sort(key=lambda item: (-item[1], item[2], str(item[0])))
        return [path for path, _score, _depth in matches]

    def find_planos_folder(self, dossier_folder: Path) -> Path | None:
        return self._find_unique_child_match(dossier_folder, "Planos")[0]

    def serie_exists_in_planos(self, planos_folder: Path, serie: str) -> bool:
        return bool(self.find_series_matches_in_planos(planos_folder, serie))

    def find_series_matches_in_planos(self, planos_folder: Path, serie: str) -> list[Path]:
        if not planos_folder.exists():
            return []
        target_key = self._normalize_series_key(serie)
        if not target_key:
            return []
        matches: list[Path] = []
        for candidate in self._iter_path_candidates(planos_folder, max_depth=5, include_files=True):
            if target_key in self._normalize_series_key(candidate.name):
                matches.append(candidate)
        return matches

    def resolve_correct_cp_folder(self, root_path: Path, cp: str, serie: str) -> Path | None:
        runtime_index = self._build_runtime_index(root_path)
        target_key = self._normalize_series_key(serie)
        for candidate in self._match_cp_candidates(runtime_index, cp):
            if self._series_found_in_candidate(candidate, target_key):
                return candidate.cp_folder
        return None

    def find_folder_5(self, dossier_folder: Path) -> Path | None:
        return self._find_phase2_folder(dossier_folder, ("5 Procedimiento de fabricación", "5 Procedimiento de fabricacion", "Procedimiento de fabricación", "Procedimiento de fabricacion", "5", "05"))

    def find_folder_6(self, dossier_folder: Path) -> Path | None:
        return self._find_phase2_folder(dossier_folder, ("6 Trazabilidad", "Trazabilidad", "6 Registros de informe de inspección", "6 Registros de informe de inspeccion", "6 Registros Informes de Inspección", "6 Registros Informes de Inspeccion", "Registros de informe de inspección", "Registros de informe de inspeccion", "Registros Informes de Inspección", "Registros Informes de Inspeccion", "6", "06"))

    def find_folder_7(self, dossier_folder: Path) -> Path | None:
        return self._find_phase2_folder(dossier_folder, ("7 Ensayos", "Ensayos", "7 Pruebas Eléctricas", "7 Pruebas Electricas", "Pruebas Eléctricas", "Pruebas Electricas", "7", "07"))

    def inspect_workbook(self, config: DossierConfig) -> DossierWorkbookInfo:
        workbook = load_workbook(config.excel_path, read_only=True, data_only=True)
        try:
            worksheet = self._resolve_worksheet(workbook, config)
            header_row = self._detect_header_row(worksheet.iter_rows(values_only=True), config)
            columns = self._extract_columns(worksheet, header_row)
            cp_column = self._resolve_column(columns, config.cp_synonyms)
            serie_column = self._resolve_column(columns, config.serie_synonyms)
            return DossierWorkbookInfo(path=config.excel_path, sheet_name=worksheet.title, header_row=header_row, columns=columns, cp_column=cp_column, serie_column=serie_column)
        finally:
            workbook.close()

    def load_rows(self, config: DossierConfig) -> tuple[DossierWorkbookInfo, list[DossierRow]]:
        workbook = load_workbook(config.excel_path, read_only=True, data_only=True)
        try:
            worksheet = self._resolve_worksheet(workbook, config)
            header_row = self._detect_header_row(worksheet.iter_rows(values_only=True), config)
            columns = self._extract_columns(worksheet, header_row)
            cp_column = self._resolve_column(columns, config.cp_synonyms)
            serie_column = self._resolve_column(columns, config.serie_synonyms)
            workbook_info = DossierWorkbookInfo(path=config.excel_path, sheet_name=worksheet.title, header_row=header_row, columns=columns, cp_column=cp_column, serie_column=serie_column)
            cp_index = self._resolve_column_index(columns, cp_column)
            serie_index = self._resolve_column_index(columns, serie_column)
            rows: list[DossierRow] = []
            seen_pair: Counter[str] = Counter()
            for row_number in range(header_row + 1, worksheet.max_row + 1):
                cp = normalize_series(worksheet.cell(row=row_number, column=cp_index).value)
                serie = normalize_series(worksheet.cell(row=row_number, column=serie_index).value)
                normalized_cp = normalize_for_match(cp)
                normalized_serie = normalize_for_match(serie)
                observation = "Ready"
                status = DossierStatus.VALID
                if not cp or not serie:
                    status = DossierStatus.ERROR
                    observation = "Missing CP or Serie"
                else:
                    seen_pair[f"{normalized_cp}::{normalized_serie}"] += 1
                    if seen_pair[f"{normalized_cp}::{normalized_serie}"] > 1:
                        status = DossierStatus.SKIPPED
                        observation = "Duplicate CP and Serie pair"
                rows.append(DossierRow(row_number=row_number, cp=cp, serie=serie, normalized_cp=normalized_cp, normalized_serie=normalized_serie, status=status, observation=observation))
            return workbook_info, rows
        finally:
            workbook.close()

    def build_validation_tree(self, config: DossierConfig, rows: list[DossierRow]) -> list[DossierValidationTreeRow]:
        runtime_index = self._build_runtime_index(config.root_path, requested_cps=[row.cp for row in rows])
        tree_rows: list[DossierValidationTreeRow] = []
        for row in rows:
            selected_cp_folder, match_state, candidates = self._inspect_row_candidates(config, row, runtime_index=runtime_index)
            tree_rows.append(DossierValidationTreeRow(row_number=row.row_number, cp=row.cp, serie=row.serie, selected_cp_folder=str(selected_cp_folder) if selected_cp_folder else "", match_state=match_state, status=row.status, observation=row.observation, candidates=candidates))
        return tree_rows

    def validate_paths(self, config: DossierConfig, rows: list[DossierRow], progress_callback=None) -> list[DossierRow]:
        total = len(rows)
        root_candidates = list(self._iter_directory_candidates(config.root_path, max_depth=1))
        index_total = max(len(root_candidates), 1)
        overall_total = max(index_total + total, 1)
        if progress_callback:
            progress_callback(0, overall_total, "Indexando carpetas CP...")

        def index_progress(current: int, total_candidates: int, message: str) -> None:
            if progress_callback:
                progress_callback(current, overall_total, message)

        runtime_index = self._build_runtime_index(config.root_path, requested_cps=[row.cp for row in rows], progress_callback=index_progress, preloaded_root_candidates=root_candidates)
        inspection_cache: dict[tuple[str, str], tuple[Path | None, str, list[DossierValidationCandidate]]] = {}

        for index, row in enumerate(rows, start=1):
            if row.status in {DossierStatus.ERROR, DossierStatus.SKIPPED}:
                if progress_callback:
                    progress_callback(index_total + index, overall_total, f"Fila {row.row_number} validada")
                continue

            cache_key = (row.normalized_cp, self._normalize_series_key(row.serie))
            if cache_key not in inspection_cache:
                inspection_cache[cache_key] = self._inspect_row_candidates(config, row, runtime_index=runtime_index)
            cp_folder_match, _cp_match_state, candidate_details = inspection_cache[cache_key]
            valid_candidates = [candidate for candidate in candidate_details if candidate.valid_for_distribution]

            if cp_folder_match is None or not valid_candidates:
                row.cp_folder = ""
                row.dossier_folder = ""
                row.planos_folder = ""
                row.folder_5 = ""
                row.folder_6 = ""
                row.folder_7 = ""
                row.series_in_planos = False
                row.matched_dossier_folders = []
                row.matched_planos_folders = []
                row.matched_folder_5 = []
                row.matched_folder_6 = []
                row.matched_folder_7 = []
                if not candidate_details:
                    row.status = DossierStatus.SKIPPED
                    row.observation = f"Carpeta CP no encontrada en {config.root_path}"
                elif any(candidate.planos_exists or candidate.folder_6_exists for candidate in candidate_details):
                    row.status = DossierStatus.SKIPPED
                    row.observation = "Serie no encontrada en 06_DOSSIER para esta CP"
                elif any(candidate.dossier_exists for candidate in candidate_details):
                    row.status = DossierStatus.SKIPPED
                    row.observation = "Existe 06_DOSSIER, pero no se encontró carpeta Planos"
                else:
                    row.status = DossierStatus.SKIPPED
                    row.observation = "Carpeta CP encontrada, pero no existe 06_DOSSIER"
                if progress_callback:
                    progress_callback(index_total + index, overall_total, f"Fila {row.row_number} validada")
                continue

            primary_candidate = valid_candidates[0]
            row.cp_folder = primary_candidate.cp_folder
            row.dossier_folder = primary_candidate.dossier_folder
            row.planos_folder = primary_candidate.planos_folder
            row.folder_5 = primary_candidate.folder_5
            row.folder_6 = primary_candidate.folder_6
            row.folder_7 = primary_candidate.folder_7
            row.series_in_planos = True
            row.matched_dossier_folders = [candidate.dossier_folder for candidate in valid_candidates if candidate.dossier_folder]
            row.matched_planos_folders = [candidate.planos_folder for candidate in valid_candidates if candidate.planos_folder]
            row.matched_folder_5 = [candidate.folder_5 for candidate in valid_candidates]
            row.matched_folder_6 = [candidate.folder_6 for candidate in valid_candidates]
            row.matched_folder_7 = [candidate.folder_7 for candidate in valid_candidates]
            row.status = DossierStatus.VALID
            row.observation = "Validada" if len(valid_candidates) == 1 else f"Validada en {len(valid_candidates)} carpetas"
            if progress_callback:
                progress_callback(index_total + index, overall_total, f"Fila {row.row_number} validada")

        return rows

    @staticmethod
    def _detect_header_row(rows: Iterable[tuple], config: DossierConfig) -> int:
        hints = {normalize_for_match(item) for item in (*config.cp_synonyms, *config.serie_synonyms)}
        best_row = 1
        best_score = float("-inf")
        for index, row in enumerate(rows, start=1):
            if index > 25:
                break
            values = [normalize_series(value) for value in row]
            non_empty = [value for value in values if value]
            if not non_empty:
                continue
            score = len(non_empty)
            normalized_row = [normalize_for_match(value) for value in non_empty]
            for candidate in normalized_row:
                if candidate in hints:
                    score += 25
                else:
                    score += max((similarity(candidate, hint) for hint in hints), default=0.0) * 5
            if any(value.isdigit() for value in non_empty):
                score -= 2
            if score > best_score:
                best_score = score
                best_row = index
        return best_row

    @staticmethod
    def _extract_columns(worksheet, header_row: int) -> list[DossierColumnInfo]:
        header_values = next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True), ())
        columns: list[DossierColumnInfo] = []
        for index, value in enumerate(header_values, start=1):
            columns.append(DossierColumnInfo(index=index, name=normalize_series(value) or f"Column {index}"))
        return columns

    @staticmethod
    def _resolve_column(columns: list[DossierColumnInfo], synonyms: tuple[str, ...]) -> str:
        best_name = ""
        best_score = -1.0
        for column in columns:
            candidate = column.name
            for synonym in synonyms:
                score = similarity(candidate, synonym)
                normalized = normalize_for_match(candidate)
                if normalized == normalize_for_match(synonym):
                    return candidate
                if score > best_score:
                    best_score = score
                    best_name = candidate
        if best_name and best_score >= 0.45:
            return best_name
        raise DossierWorkbookHeaderError(f"No se encontró una columna que coincida con: {', '.join(synonyms)}")

    @staticmethod
    def _resolve_column_index(columns: list[DossierColumnInfo], column_name: str) -> int:
        target = normalize_for_match(column_name)
        for column in columns:
            if normalize_for_match(column.name) == target:
                return column.index
        raise DossierWorkbookHeaderError(f"La columna '{column_name}' no se encontró en el libro.")

    def _inspect_row_candidates(self, config: DossierConfig, row: DossierRow, runtime_index: DossierRuntimeIndex | None = None) -> tuple[Path | None, str, list[DossierValidationCandidate]]:
        runtime_index = runtime_index or self._build_runtime_index(config.root_path)
        candidate_entries = self._match_cp_candidates(runtime_index, row.cp)
        candidates: list[DossierValidationCandidate] = []
        valid_paths: list[Path] = []
        target_series_key = self._normalize_series_key(row.serie)

        for entry in candidate_entries:
            series_found = self._series_found_in_candidate(entry, target_series_key)
            valid_for_distribution = bool(entry.dossier_folder and series_found)
            if valid_for_distribution:
                valid_paths.append(entry.cp_folder)
            candidates.append(
                DossierValidationCandidate(
                    cp_folder=str(entry.cp_folder),
                    dossier_folder=str(entry.dossier_folder) if entry.dossier_folder else "",
                    planos_folder=str(entry.planos_folder) if entry.planos_folder else "",
                    dossier_exists=entry.dossier_folder is not None,
                    planos_exists=entry.planos_folder is not None,
                    series_found=series_found,
                    folder_5_exists=entry.folder_5 is not None,
                    folder_6_exists=entry.folder_6 is not None,
                    folder_7_exists=entry.folder_7 is not None,
                    selected=valid_for_distribution,
                    valid_for_distribution=valid_for_distribution,
                    reason=self._candidate_reason(entry.dossier_folder, entry.planos_folder, series_found, valid_for_distribution),
                    folder_5=str(entry.folder_5) if entry.folder_5 else "",
                    folder_6=str(entry.folder_6) if entry.folder_6 else "",
                    folder_7=str(entry.folder_7) if entry.folder_7 else "",
                    series_locations=[str(path) for path in self._find_series_locations(entry, target_series_key)],
                )
            )

        if valid_paths:
            return valid_paths[0], "matched", candidates
        if not candidate_entries:
            return None, "missing", []
        if len(candidate_entries) > 1:
            return None, "ambiguous", candidates
        return None, "missing-series", candidates

    @staticmethod
    def _candidate_reason(dossier_folder: Path | None, planos_folder: Path | None, series_found: bool, valid_for_distribution: bool) -> str:
        if dossier_folder is None:
            return "Carpeta candidata encontrada, pero no existe 06_DOSSIER."
        if planos_folder is None:
            return "Existe 06_DOSSIER, pero no se encontró carpeta Planos."
        if not series_found:
            return "La serie no se encontró en esta carpeta CP."
        if valid_for_distribution:
            return "Carpeta válida para distribución."
        return "Candidata no válida para distribución."

    def _find_phase2_folder(self, dossier_folder: Path, terms: tuple[str, ...]) -> Path | None:
        if not dossier_folder.is_dir():
            return None
        numbered_terms = {self._extract_leading_number(term) for term in terms if self._extract_leading_number(term)}
        try:
            candidates = [item for item in dossier_folder.iterdir() if item.is_dir()]
        except OSError:
            return None
        if numbered_terms:
            numbered_candidates = [candidate for candidate in candidates if self._extract_leading_number(candidate.name) in numbered_terms]
            if numbered_candidates:
                candidates = numbered_candidates
        best_path: Path | None = None
        best_score = 0.0
        for candidate in candidates:
            candidate_key = normalize_for_match(candidate.name)
            candidate_number = self._extract_leading_number(candidate.name)
            for term in terms:
                term_key = normalize_for_match(term)
                term_number = self._extract_leading_number(term)
                if term_number and candidate_number and term_number != candidate_number:
                    continue
                score = similarity(candidate.name, term_key)
                if candidate_key == term_key:
                    return candidate
                if term_key in candidate_key:
                    score += 0.35
                if term_number and candidate_number == term_number:
                    score += 0.4
                if score > best_score:
                    best_score = score
                    best_path = candidate
        return best_path if best_score >= 0.45 else None

    @staticmethod
    def _extract_leading_number(value: str) -> str:
        match = re.match(r"\s*(\d+)", normalize_series(value))
        return match.group(1) if match else ""

    @staticmethod
    def _extract_cp_code(value: str) -> str:
        normalized_raw = normalize_series(value).strip()
        if normalized_raw.isdigit():
            return normalized_raw
        normalized_value = normalize_for_match(value)
        match = _CP_IDENTIFIER_PATTERN.search(normalized_value)
        if match:
            return match.group("code")
        leading_digits = re.match(r"^\s*(\d{5,})", normalized_raw)
        return leading_digits.group(1) if leading_digits else ""

    @staticmethod
    def _resolve_worksheet(workbook, config: DossierConfig):
        if config.sheet_name and config.sheet_name in workbook.sheetnames:
            return workbook[config.sheet_name]
        return workbook[workbook.sheetnames[0]]

    def _find_unique_child_match(self, parent: Path, target: str) -> tuple[Path | None, str]:
        if not parent.is_dir():
            return None, "missing"
        best_paths: list[Path] = []
        best_score = 0.0
        try:
            candidates = [item for item in parent.iterdir() if item.is_dir()]
        except OSError:
            return None, "missing"
        for candidate in candidates:
            score = self._score_directory_name(candidate.name, target)
            if score < 0.45:
                continue
            if score > best_score + 1e-9:
                best_score = score
                best_paths = [candidate]
            elif abs(score - best_score) <= 1e-9:
                best_paths.append(candidate)
        if not best_paths:
            return None, "missing"
        if len(best_paths) > 1:
            return None, "ambiguous"
        return best_paths[0], ""

    def _iter_directory_candidates(self, root_path: Path, max_depth: int = 2):
        stack = [(root_path, 0)]
        while stack:
            current, depth = stack.pop()
            if depth > max_depth:
                continue
            if current.is_dir() and current != root_path:
                yield current
            if depth == max_depth:
                continue
            try:
                children = list(current.iterdir())
            except OSError:
                continue
            for child in children:
                try:
                    if child.is_dir():
                        stack.append((child, depth + 1))
                except OSError:
                    continue

    def _iter_path_candidates(self, root_path: Path, max_depth: int = 5, include_files: bool = False):
        stack = [(root_path, 0)]
        while stack:
            current, depth = stack.pop()
            if depth > max_depth:
                continue
            if current != root_path and (current.is_dir() or (include_files and current.is_file())):
                yield current
            if depth == max_depth or not current.is_dir():
                continue
            try:
                children = list(current.iterdir())
            except OSError:
                continue
            for child in children:
                try:
                    is_directory = child.is_dir()
                except OSError:
                    continue
                if is_directory or include_files:
                    stack.append((child, depth + 1))

    def _build_runtime_index(self, root_path: Path, requested_cps: Iterable[str] | None = None, progress_callback=None, preloaded_root_candidates: list[Path] | None = None) -> DossierRuntimeIndex:
        candidates: list[DossierIndexedCandidate] = []
        cp_code_lookup: dict[str, list[DossierIndexedCandidate]] = defaultdict(list)

        requested_codes: set[str] = set()
        requested_names: list[str] = []
        if requested_cps is not None:
            for cp in requested_cps:
                cp_code = self._extract_cp_code(cp)
                if cp_code:
                    requested_codes.add(cp_code)
                else:
                    requested_names.append(normalize_for_match(cp))

        root_candidates = preloaded_root_candidates if preloaded_root_candidates is not None else list(self._iter_directory_candidates(root_path, max_depth=1))
        sorted_candidates = sorted(root_candidates, key=lambda item: len(str(item)))
        total_candidates = max(len(sorted_candidates), 1)

        for scan_index, cp_folder in enumerate(sorted_candidates, start=1):
            if progress_callback:
                progress_callback(scan_index, total_candidates, f"Indexando CP: {cp_folder.name}")
            cp_name_key = normalize_for_match(cp_folder.name)
            cp_code = self._extract_cp_code(cp_folder.name)

            if requested_cps is not None:
                matches_requested = False
                if cp_code and cp_code in requested_codes:
                    matches_requested = True
                elif any(name_key == cp_name_key or name_key in cp_name_key or cp_name_key in name_key for name_key in requested_names):
                    matches_requested = True
                if not matches_requested:
                    continue

            dossier_folders = self.find_dossier_folders(cp_folder)
            if not dossier_folders:
                indexed = DossierIndexedCandidate(
                    cp_folder=cp_folder,
                    cp_name_key=cp_name_key,
                    cp_code=cp_code,
                    dossier_folder=None,
                    planos_folder=None,
                    folder_5=None,
                    folder_6=None,
                    folder_7=None,
                    series_keys=(),
                )
                candidates.append(indexed)
                if indexed.cp_code:
                    cp_code_lookup[indexed.cp_code].append(indexed)
                continue

            for dossier_folder in dossier_folders:
                planos_folder = self.find_planos_folder(dossier_folder)
                folder_5 = self.find_folder_5(dossier_folder)
                folder_6 = self.find_folder_6(dossier_folder)
                folder_7 = self.find_folder_7(dossier_folder)
                series_keys = self._collect_series_keys(dossier_folder)
                indexed = DossierIndexedCandidate(
                    cp_folder=cp_folder,
                    cp_name_key=cp_name_key,
                    cp_code=cp_code,
                    dossier_folder=dossier_folder,
                    planos_folder=planos_folder,
                    folder_5=folder_5,
                    folder_6=folder_6,
                    folder_7=folder_7,
                    series_keys=tuple(series_keys),
                )
                candidates.append(indexed)
                if indexed.cp_code:
                    cp_code_lookup[indexed.cp_code].append(indexed)

        return DossierRuntimeIndex(candidates=tuple(candidates), cp_code_lookup={key: tuple(value) for key, value in cp_code_lookup.items()})

    def _match_cp_candidates(self, runtime_index: DossierRuntimeIndex, cp: str) -> list[DossierIndexedCandidate]:
        target_cp_code = self._extract_cp_code(cp)
        target_key = normalize_for_match(cp)
        if target_cp_code:
            return list(runtime_index.cp_code_lookup.get(target_cp_code, ()))

        matches: list[DossierIndexedCandidate] = []
        for candidate in runtime_index.candidates:
            if candidate.cp_name_key == target_key or target_key in candidate.cp_name_key or similarity(candidate.cp_folder.name, target_key) >= 0.55:
                matches.append(candidate)
        return sorted(matches, key=lambda item: len(str(item.cp_folder)))

    def _series_found_in_candidate(self, candidate: DossierIndexedCandidate, target_series_key: str) -> bool:
        if not target_series_key:
            return False
        return any(target_series_key in series_key for series_key in candidate.series_keys)

    def _find_series_locations(self, candidate: DossierIndexedCandidate, target_series_key: str) -> list[Path]:
        if not target_series_key or not candidate.dossier_folder:
            return []
        matches: list[Path] = []
        for root, dirs, files in self._safe_walk(candidate.dossier_folder):
            for directory_name in dirs:
                if target_series_key in self._normalize_series_key(directory_name):
                    matches.append(root / directory_name)
            for file_name in files:
                if target_series_key in self._normalize_series_key(file_name):
                    matches.append(root / file_name)
        return matches

    def _looks_like_distribution_root(self, candidate: Path, child_dirs: list[str]) -> bool:
        if not child_dirs:
            return False
        has_planos = False
        has_distribution_folder = False
        for child_name in child_dirs:
            normalized = normalize_for_match(child_name)
            if "planos" in normalized:
                has_planos = True
            if normalized.startswith("5") or normalized.startswith("6") or normalized.startswith("7"):
                has_distribution_folder = True
        return has_planos and has_distribution_folder

    def _collect_series_keys(self, dossier_folder: Path) -> tuple[str, ...]:
        seen_series: set[str] = set()
        series_keys: list[str] = []
        for root, dirs, files in self._safe_walk(dossier_folder):
            for directory_name in dirs:
                series_key = self._normalize_series_key(directory_name)
                if series_key and series_key not in seen_series:
                    seen_series.add(series_key)
                    series_keys.append(series_key)
            for file_name in files:
                series_key = self._normalize_series_key(file_name)
                if series_key and series_key not in seen_series:
                    seen_series.add(series_key)
                    series_keys.append(series_key)
        return tuple(series_keys)

    def _safe_walk(self, root_path: Path):
        if not root_path.is_dir():
            return
        try:
            walker = os.walk(root_path, topdown=True, onerror=lambda _error: None)
        except OSError:
            return
        for current_root, dirs, files in walker:
            current_path = Path(current_root)
            filtered_dirs: list[str] = []
            for directory_name in list(dirs):
                if directory_name.lower() == '_backups':
                    continue
                try:
                    candidate = current_path / directory_name
                    if candidate.is_dir():
                        filtered_dirs.append(directory_name)
                except OSError:
                    continue
            dirs[:] = filtered_dirs
            yield current_path, list(dirs), list(files)

    @staticmethod
    def _score_directory_name(candidate_name: str, target: str) -> float:
        target_key = normalize_for_match(target)
        candidate_key = normalize_for_match(candidate_name)
        score = similarity(candidate_name, target_key)
        if target_key in candidate_key:
            score += 0.5
        target_number = DossierValidatorService._extract_leading_number(target)
        candidate_number = DossierValidatorService._extract_leading_number(candidate_name)
        if target_number and candidate_number == target_number:
            score += 0.25
        return score

    @staticmethod
    def _normalize_series_key(value: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", normalize_for_match(value))
